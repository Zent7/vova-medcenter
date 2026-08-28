from base64 import b64encode
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import sys
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session


sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import app.models  # noqa: E402,F401
from app.api.v1.routes.imports import (  # noqa: E402
    build_client_payload,
    commit_client_excel_import,
    find_existing_client_for_import,
    parse_optional_date,
    preview_client_excel_import,
    merge_registration_text,
    read_client_excel_rows,
    resolve_import_service,
    rows_to_client_records,
    build_service_lookup,
    service_import_label,
    validate_client_import_rows,
)
from app.db.base import Base  # noqa: E402
from app.models.center import Center  # noqa: E402
from app.models.client import Client  # noqa: E402
from app.models.encounter import Encounter  # noqa: E402
from app.models.encounter_service import EncounterService  # noqa: E402
from app.models.payment import Payment  # noqa: E402
from app.models.service import Service  # noqa: E402
from app.models.user import Role, User  # noqa: E402
from app.schemas.imports import ClientImportExcelRequest  # noqa: E402


class ClientExcelImportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)

        center = Center(id=1, code="center", name="Медцентр", is_active=True)
        role = Role(id=1, code="admin", name="Администратор")
        user = User(
            id=1,
            center_id=1,
            role_id=1,
            login="admin",
            password_hash="test",
            full_name="Администратор",
            is_active=True,
        )
        self.service = Service(
            id=1,
            code="service-pool",
            name="Справка в бассейн",
            price=Decimal("1000.00"),
            is_active=True,
        )
        self.db.add_all([center, role, user, self.service])
        self.db.commit()
        self.request = ClientImportExcelRequest(file_name="clients.xlsx", file_content_base64="")

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def _row(self, **overrides):
        row = {
            "row_number": 2,
            "patient_number": None,
            "last_name": "Иванов",
            "first_name": "Иван",
            "middle_name": "Иванович",
            "birth_date": date(1990, 5, 20),
            "sex": "муж",
            "phone": "+79990000000",
            "snils": "111-222-333 44",
            "registration_text": "Санкт-Петербург",
            "organization": "ООО Тест",
            "service": service_import_label(self.service),
            "encounter_date": date(2026, 7, 29),
            "notes": "Тестовый импорт",
            "_birth_date_invalid": False,
            "_document_issued_date_invalid": False,
            "_encounter_date_invalid": False,
        }
        row.update(overrides)
        return row

    def test_parses_native_excel_date_serial_and_iso_datetime(self):
        self.assertEqual(parse_optional_date("43971"), date(2020, 5, 20))
        self.assertEqual(parse_optional_date("2026-07-29T12:30:00"), date(2026, 7, 29))

    def test_empty_file_and_missing_birth_date_have_clear_errors(self):
        with self.assertRaisesRegex(ValueError, "не найдено ни одной"):
            validate_client_import_rows([])

        with self.assertRaisesRegex(ValueError, "не заполнена дата рождения"):
            validate_client_import_rows([self._row(birth_date=None)])

    def test_commit_creates_client_encounter_service_and_payment(self):
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=[self._row()],
        ):
            result = commit_client_excel_import(self.request, db=self.db)

        self.assertEqual(result.created, 1)
        self.assertEqual(result.updated, 0)
        self.assertEqual(result.encounters_created, 1)
        self.assertEqual(self.db.scalar(select(func.count(Client.id))), 1)
        self.assertEqual(self.db.scalar(select(func.count(Encounter.id))), 1)
        self.assertEqual(self.db.scalar(select(func.count(EncounterService.id))), 1)
        self.assertEqual(self.db.scalar(select(func.count(Payment.id))), 1)

        client = self.db.execute(select(Client)).scalar_one()
        encounter = self.db.execute(select(Encounter)).scalar_one()
        encounter_service = self.db.execute(select(EncounterService)).scalar_one()
        payment = self.db.execute(select(Payment)).scalar_one()
        self.assertEqual(client.birth_date, date(1990, 5, 20))
        self.assertEqual(client.legacy_payload_json["birth_date"], "1990-05-20")
        self.assertEqual(client.encounter_date_text, "2026-07-29")
        self.assertEqual(encounter.client_id, client.id)
        self.assertEqual(encounter_service.service_id, self.service.id)
        self.assertEqual(payment.amount, Decimal("1000.00"))

    def test_unknown_service_warns_but_still_imports_client(self):
        row = self._row(service="Несуществующая услуга")
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=[row],
        ):
            preview = preview_client_excel_import(self.request, db=self.db)
            result = commit_client_excel_import(self.request, db=self.db)

        self.assertEqual(preview.service_rows, 0)
        self.assertIn("не найдена", " ".join(preview.service_warnings))
        self.assertEqual(result.created, 1)
        self.assertEqual(result.encounters_created, 0)
        self.assertIn("не найдена", " ".join(result.service_warnings))
        self.assertEqual(self.db.scalar(select(func.count(Client.id))), 1)

    def test_deleted_client_is_not_matched_and_stays_deleted(self):
        deleted = Client(
            patient_number=9,
            last_name="Удалённый",
            first_name="Клиент",
            birth_date=date(1990, 5, 20),
            snils="999-888-777 66",
            deleted_at=datetime(2026, 1, 1),
        )
        self.db.add(deleted)
        self.db.commit()

        row = self._row(snils="999-888-777 66", service=None)
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=[row],
        ):
            result = commit_client_excel_import(self.request, db=self.db)

        self.assertEqual(result.created, 1)
        self.assertEqual(result.updated, 0)
        self.db.refresh(deleted)
        self.assertIsNotNone(deleted.deleted_at)
        self.assertEqual(deleted.last_name, "Удалённый")

    def test_same_snils_with_different_birth_date_creates_a_new_client(self):
        existing = Client(
            patient_number=9,
            last_name="Иванов",
            first_name="Иван",
            middle_name="Иванович",
            birth_date=date(1990, 5, 20),
            snils="999-888-777 66",
        )
        self.db.add(existing)
        self.db.commit()

        row = self._row(
            snils="999-888-777 66",
            birth_date=date(1991, 5, 20),
            service=None,
        )
        self.assertEqual(find_existing_client_for_import(self.db, row), (None, None))
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=[row],
        ):
            result = commit_client_excel_import(self.request, db=self.db)

        self.assertEqual(result.created, 1)
        self.assertEqual(result.updated, 0)
        self.assertEqual(self.db.scalar(select(func.count(Client.id))), 2)

    def test_service_warning_rows_counts_every_row_not_only_shown(self):
        rows = [
            self._row(row_number=number, service="Несуществующая услуга")
            for number in range(2, 27)
        ]
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=rows,
        ):
            preview = preview_client_excel_import(self.request, db=self.db)

        self.assertEqual(preview.service_warning_rows, 25)
        self.assertEqual(len(preview.service_warnings), 20)

    def test_update_keeps_fields_missing_from_template(self):
        existing = Client(
            patient_number=7,
            last_name="Иванов",
            first_name="Иван",
            middle_name="Иванович",
            birth_date=date(1990, 5, 20),
            snils="111-222-333 44",
            document_type="Паспорт РФ",
            document_series="4501",
            document_number="123456",
            document_issued_by="ГУ МВД России",
            email="ivanov@example.com",
            address_text="Тверь, Лесная 1",
            admission_category="Медкомиссия",
            reference_number="МК-001",
        )
        self.db.add(existing)
        self.db.commit()

        # Шаблон для завода не содержит паспортных колонок вовсе.
        narrow_row = {
            "row_number": 2,
            "patient_number": None,
            "last_name": "Иванов",
            "first_name": "Иван",
            "middle_name": "Иванович",
            "birth_date": date(1990, 5, 20),
            "snils": "111-222-333 44",
            "organization": "ООО Новый Завод",
            "_birth_date_invalid": False,
        }
        with patch(
            "app.api.v1.routes.imports.read_client_excel_rows",
            return_value=[narrow_row],
        ):
            result = commit_client_excel_import(self.request, db=self.db)

        self.assertEqual(result.updated, 1)
        self.assertEqual(result.created, 0)
        client = self.db.execute(select(Client)).scalar_one()
        self.assertEqual(client.organization, "ООО Новый Завод")
        self.assertEqual(client.document_series, "4501")
        self.assertEqual(client.document_number, "123456")
        self.assertEqual(client.document_type, "Паспорт РФ")
        self.assertEqual(client.email, "ivanov@example.com")
        self.assertEqual(client.address_text, "Тверь, Лесная 1")
        self.assertEqual(client.admission_category, "Медкомиссия")
        self.assertEqual(client.reference_number, "МК-001")


class FactoryTemplateImportTests(unittest.TestCase):
    """Разбор файла, собранного по выдаваемому заводу шаблону."""

    TEMPLATE_PATH = (
        Path(__file__).resolve().parents[2]
        / "frontend"
        / "public"
        / "demo"
        / "client-import-template.xlsx"
    )

    def _request_from_rows(self, rows, extra=None):
        workbook = load_workbook(self.TEMPLATE_PATH)
        sheet = workbook["Clients"]
        for row_offset, values in enumerate(rows, start=2):
            for column_offset, value in enumerate(values, start=1):
                sheet.cell(row=row_offset, column=column_offset, value=value)
        for cell_reference, value, number_format in extra or []:
            cell = sheet[cell_reference]
            cell.value = value
            cell.number_format = number_format
        buffer = BytesIO()
        workbook.save(buffer)
        return ClientImportExcelRequest(
            file_name="clients.xlsx",
            file_content_base64=b64encode(buffer.getvalue()).decode(),
        )

    def test_reads_every_column_of_the_factory_template(self):
        request = self._request_from_rows(
            [
                [
                    "ЛМК", "Иванов", "Иван", "Иванович", "15.04.1987", "муж",
                    "Тверская", "Тверь", "Лесная", "1", "2", "15",
                    "Завод Пример", "Слесарь", "Шум, вибрация", "111-222-333 44",
                    "12.03.2025", "от завода",
                ]
            ]
        )
        rows = read_client_excel_rows(request)
        validate_client_import_rows(rows)
        self.assertEqual(len(rows), 1)

        payload = build_client_payload(rows[0], 1)
        self.assertEqual(payload["last_name"], "Иванов")
        self.assertEqual(payload["birth_date"], date(1987, 4, 15))
        self.assertEqual(
            payload["registration_text"],
            "Тверская, г. Тверь, ул. Лесная, д. 1, корп. 2, кв. 15",
        )
        self.assertEqual(payload["organization"], "Завод Пример")
        self.assertEqual(payload["profession"], "Слесарь")
        self.assertEqual(payload["indications"], "Шум, вибрация")
        self.assertEqual(payload["flg"], "12.03.2025")
        self.assertEqual(rows[0]["service"], "ЛМК")

    def test_date_typed_as_real_excel_date_is_not_read_as_number(self):
        request = self._request_from_rows(
            [["ЛМК", "Датов", "Пётр", "", "", "муж"]],
            extra=[
                ("E2", datetime(1993, 6, 5), "DD.MM.YYYY"),
                ("Q2", datetime(2024, 11, 2), "DD.MM.YYYY"),
            ],
        )
        rows = read_client_excel_rows(request)
        validate_client_import_rows(rows)
        self.assertEqual(rows[0]["birth_date"], date(1993, 6, 5))
        self.assertEqual(build_client_payload(rows[0], 1)["flg"], "02.11.2024")

    def test_header_hint_in_parentheses_does_not_break_the_column(self):
        workbook = load_workbook(self.TEMPLATE_PATH)
        self.assertEqual(
            workbook["Clients"]["E1"].value, "Дата рождения (формат 31.12.2026)"
        )

    def test_two_word_toponyms_still_get_their_prefix(self):
        merged = merge_registration_text(
            {
                "reg_region": "Свердловская",
                "reg_city": "Верхняя Пышма",
                "reg_street": "Большая Морская",
                "reg_house": "12",
            }
        )
        self.assertEqual(
            merged, "Свердловская, г. Верхняя Пышма, ул. Большая Морская, д. 12"
        )

    def test_prefix_written_by_hand_is_not_duplicated(self):
        merged = merge_registration_text(
            {"reg_city": "г. Тверь", "reg_street": "улица Мира", "reg_house": "д. 5"}
        )
        self.assertEqual(merged, "г. Тверь, улица Мира, д. 5")

    def test_legend_rows_under_the_table_are_skipped(self):
        request = self._request_from_rows(
            [
                ["ЛМК", "Иванов", "Иван", "", "15.04.1987", "муж"],
                ["Проф"],
                ["Тракторная"],
            ]
        )
        rows = read_client_excel_rows(request)
        validate_client_import_rows(rows)
        self.assertEqual([row["last_name"] for row in rows], ["Иванов"])

    def test_row_with_data_but_without_a_name_is_reported_not_dropped(self):
        request = self._request_from_rows(
            [["ЛМК", "", "", "", "", "муж", "", "", "", "", "", "", "", "", "", "111-222-333 44"]]
        )
        rows = read_client_excel_rows(request)
        self.assertEqual(len(rows), 1)
        with self.assertRaises(ValueError) as error:
            validate_client_import_rows(rows)
        self.assertIn("не заполнена фамилия", str(error.exception))


class ServiceValueAliasTests(unittest.TestCase):
    """Колонка «Тип документа» — сокращения услуг из шаблона."""

    def setUp(self):
        self.services = [
            Service(id=1, code="s1", name="ЛМК", price=Decimal("4000.00"), is_active=True),
            Service(id=2, code="s2", name="Продление ЛМК", price=Decimal("3500.00"), is_active=True),
            Service(id=3, code="s3", name="ГИМС", price=Decimal("3500.00"), is_active=True),
            Service(id=4, code="s4", name="Профосмотр", price=Decimal("3500.00"), is_active=True),
            Service(
                id=5,
                legacy_source_id=8,
                code="s5",
                name="Медицинская комиссия",
                price=Decimal("4000.00"),
                is_active=True,
            ),
            Service(
                id=6,
                legacy_source_id=29,
                code="s6",
                name="Медицинская комиссия",
                price=Decimal("3500.00"),
                is_active=True,
            ),
            Service(
                id=7,
                legacy_source_id=7,
                code="s7",
                name="071У",
                price=Decimal("4000.00"),
                is_active=True,
            ),
        ]
        self.lookup = build_service_lookup(self.services)

    def _resolve(self, value):
        return resolve_import_service({"service": value, "row_number": 2}, self.lookup)

    def test_unambiguous_values_resolve_to_a_service(self):
        for value, expected in [
            ("ЛМК", "ЛМК"),
            ("Продление ЛМК", "Продление ЛМК"),
            ("ГИМС", "ГИМС"),
            ("Проф", "Профосмотр"),
            ("Водительская", "Медицинская комиссия"),
            ("Тракторная", "071У"),
        ]:
            with self.subTest(value=value):
                service, warning = self._resolve(value)
                self.assertIsNotNone(service)
                self.assertEqual(service.name, expected)
                self.assertIsNone(warning)

        driver, driver_warning = self._resolve("Водительская")
        self.assertEqual(driver.id, 5)
        self.assertIsNone(driver_warning)

    def test_ambiguous_and_unknown_values_are_left_to_the_operator(self):
        for value in ("Медицинская комиссия", "Несуществующая услуга"):
            with self.subTest(value=value):
                service, warning = self._resolve(value)
                self.assertIsNone(service)
                self.assertIn(value, warning)


class LegacyDocumentTypeColumnTests(unittest.TestCase):
    """Старые файлы, где «Тип документа» — это паспорт, а не услуга."""

    def test_separate_service_column_keeps_document_type_as_a_passport(self):
        rows = [
            {0: "Фамилия", 1: "Имя", 2: "Дата рождения", 3: "Тип документа", 4: "Услуга"},
            {0: "Иванов", 1: "Иван", 2: "15.04.1987", 3: "Паспорт РФ", 4: "ЛМК"},
        ]
        record = rows_to_client_records(rows)[0]
        self.assertEqual(record["document_type"], "Паспорт РФ")
        self.assertEqual(record["service"], "ЛМК")

    def test_blank_service_column_does_not_erase_the_document_type(self):
        rows = [
            {0: "Фамилия", 1: "Имя", 2: "Дата рождения", 3: "Тип документа", 4: "Услуга"},
            {0: "Иванов", 1: "Иван", 2: "15.04.1987", 3: "Паспорт РФ", 4: None},
        ]
        record = rows_to_client_records(rows)[0]
        self.assertEqual(record["document_type"], "Паспорт РФ")
        self.assertIsNone(record["service"])

    def test_without_a_service_column_document_type_means_the_service(self):
        rows = [
            {0: "Тип документа", 1: "Фамилия", 2: "Имя", 3: "Дата рождения"},
            {0: "ЛМК", 1: "Иванов", 2: "Иван", 3: "15.04.1987"},
        ]
        record = rows_to_client_records(rows)[0]
        self.assertEqual(record["service"], "ЛМК")
        self.assertIsNone(record.get("document_type"))


if __name__ == "__main__":
    unittest.main()
