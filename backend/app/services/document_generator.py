from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
import shutil
import struct
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from xml.sax.saxutils import escape as escape_xml_text
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import xlrd
import xlwt
from sqlalchemy import select
from sqlalchemy.orm import Session
from xlrd import xldate
from xlrd.compdoc import CompDoc
from xlutils.copy import copy as copy_xls_workbook

from app.core.config import settings
from app.models.blank_form import (
    BLANK_STATUS_FREE,
    BLANK_STATUS_ISSUED,
    BLANK_TYPE_DRIVER_MEDICAL_CERTIFICATE,
    BlankForm,
)
from app.models.client import Client
from app.models.doctor_exam import DoctorExam
from app.models.document_journal import DocumentJournalEntry
from app.models.document_template import DocumentTemplate
from app.models.encounter import Encounter
from app.models.encounter_service import EncounterService
from app.models.generated_document import GeneratedDocument
from app.models.medical_record import MedicalRecord, MedicalRecordEntry
from app.models.service import Service
from app.schemas.document_generation import DocumentGenerateResponse
from app.services.audit import write_audit_log
from app.services.blank_forms import (
    issue_specific_blank,
    issue_next_blank,
    resolve_required_blank_type,
    reuse_blank_for_existing_document,
)
from app.services.document_context import (
    MONTH_NAMES,
    _add_calendar_months,
    _split_address,
    build_document_context,
)
from app.services.new_xls_templates import (
    NEW_XLS_TEMPLATE_BY_FILE,
    NEW_XLS_TEMPLATE_BY_SHEET,
    PLACEHOLDER_LENGTH,
    NewXlsTemplateSpec,
    new_xls_placeholder,
)

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": W_NS}
ET.register_namespace("w", W_NS)

SOAP_NS = "http://schemas.xmlsoap.org/soap/envelope/"
MIAC_NS = "http://iac.spb.ru/mb#"
XMLDSIG_NS = "http://www.w3.org/2000/09/xmldsig#"
MIAC_DRIVER_TEMPLATE = "Водительская(новая).xml"
MIAC_GUARD_TEMPLATE = "Чод_новый.xml"
MIAC_GIMS_TEMPLATE = "ГИМС_шаблон_для_загрузки_из_файла.xml"

CHAIRMAN_EXAM_DATE_TEMPLATE_FILES = frozenset(
    {
        "cправкабассейн_шаблон.docx",
        "гс новый формат.xls",
        "гт.xls",
        "гто1144_шаблон.docx",
        "скк 070 новый формат.xls",
        "скк 72 новый формат.xls",
        "спорт.xls",
    }
)
CHAIRMAN_CERTIFICATE_082_TEMPLATE_FILES = frozenset({"082у_шаблон.docx"})

ET.register_namespace("soapenv", SOAP_NS)
ET.register_namespace("mb", MIAC_NS)

PROF_EXTRACT_DOCTOR_ROWS: tuple[tuple[str, str, int], ...] = (
    ("therapist", "Терапевт", 32),
    ("psychiatrist", "Психиатр", 34),
    ("psychiatrist-narcologist", "Психиатр-нарколог", 37),
    ("neurologist", "Невролог", 39),
    ("otolaryngologist", "Отоларинголог", 41),
    ("surgeon", "Хирург", 43),
    ("gynecologist", "Гинеколог", 45),
    ("ophthalmologist", "Офтальмолог", 48),
    ("dermatologist", "Дерматовенеролог", 50),
    ("dentist", "Стоматолог", 52),
)
PROF_EXTRACT_DOCTOR_COL = 44
PROF_EXTRACT_DATE_COL = 54
PROF_EXTRACT_CONCLUSION_COL = 63
PROF_EXTRACT_SEQUENCE_COL = 42
PROF_EXTRACT_CLEARED_DOCTOR_ROWS: tuple[int, ...] = ()
PROF_EXTRACT_CLIENT_DOCTOR_FIELDS = {
    "therapist": "doctor_therapist",
    "psychiatrist": "doctor_psychiatrist",
    "psychiatrist-narcologist": "doctor_psychiatrist",
    "neurologist": "doctor_neurologist",
    "otolaryngologist": "doctor_otolaryngologist",
    "surgeon": "doctor_surgeon",
    "gynecologist": "doctor_gynecologist",
    "ophthalmologist": "doctor_ophthalmologist",
    "dermatologist": "doctor_dermatologist",
    "dentist": "doctor_stomatologist",
}
PROF_AMB_EXAM_BLOCKS: tuple[dict[str, tuple[int, int]], ...] = (
    {
        "date_cell": (1, 24),
        "title_cell": (2, 10),
        "complaints_cell": (3, 10),
        "anamnesis_cell": (4, 13),
        "objective_cell": (6, 1),
        "diagnosis_cell": (9, 1),
        "doctor_cell": (14, 11),
    },
    {
        "date_cell": (15, 24),
        "title_cell": (16, 10),
        "complaints_cell": (17, 10),
        "anamnesis_cell": (18, 13),
        "objective_cell": (20, 1),
        "diagnosis_cell": (23, 1),
        "doctor_cell": (28, 11),
    },
    {
        "date_cell": (51, 24),
        "title_cell": (52, 10),
        "complaints_cell": (53, 10),
        "anamnesis_cell": (54, 13),
        "objective_cell": (56, 1),
        "diagnosis_cell": (59, 1),
        "doctor_cell": (64, 11),
    },
    {
        "date_cell": (51, 55),
        "title_cell": (52, 41),
        "complaints_cell": (53, 41),
        "anamnesis_cell": (54, 44),
        "objective_cell": (56, 32),
        "diagnosis_cell": (59, 32),
        "doctor_cell": (64, 42),
    },
    {
        "date_cell": (66, 24),
        "title_cell": (67, 10),
        "complaints_cell": (68, 10),
        "anamnesis_cell": (69, 13),
        "objective_cell": (71, 1),
        "diagnosis_cell": (74, 1),
        "doctor_cell": (79, 11),
    },
    {
        "date_cell": (66, 55),
        "title_cell": (67, 41),
        "complaints_cell": (68, 41),
        "anamnesis_cell": (69, 44),
        "objective_cell": (71, 32),
        "diagnosis_cell": (74, 32),
        "doctor_cell": (79, 42),
    },
    {
        "date_cell": (81, 24),
        "title_cell": (82, 10),
        "complaints_cell": (83, 10),
        "anamnesis_cell": (84, 13),
        "objective_cell": (86, 1),
        "diagnosis_cell": (89, 1),
        "doctor_cell": (94, 11),
    },
    {
        "date_cell": (81, 55),
        "title_cell": (82, 41),
        "complaints_cell": (83, 41),
        "anamnesis_cell": (84, 44),
        "objective_cell": (86, 32),
        "diagnosis_cell": (89, 32),
        "doctor_cell": (94, 42),
    },
    {
        "date_cell": (96, 24),
        "title_cell": (97, 10),
        "complaints_cell": (98, 10),
        "anamnesis_cell": (99, 13),
        "objective_cell": (101, 1),
        "diagnosis_cell": (104, 1),
        "doctor_cell": (109, 11),
    },
    {
        "date_cell": (96, 55),
        "title_cell": (97, 41),
        "complaints_cell": (98, 41),
        "anamnesis_cell": (99, 44),
        "objective_cell": (101, 32),
        "diagnosis_cell": (104, 32),
        "doctor_cell": (109, 42),
    },
)


def _is_contract_template(template: DocumentTemplate) -> bool:
    text = " ".join(
        [
            str(template.name or ""),
            str(template.code or ""),
            str(template.file_name or ""),
        ]
    ).lower()
    return "договор" in text or "contract" in text


def _cleanup_contract_xml(xml_text: str) -> str:
    cleanup_patterns = [
        r"Баронина\s+Виктора\s+Евгеньевича",
        r"Баронин\s+Виктор\s+Евгеньевич",
        r"П\s*о\s*д\s*п\s*и\s*с\s*ь",
    ]
    for pattern in cleanup_patterns:
        xml_text = re.sub(pattern, "", xml_text, flags=re.IGNORECASE)
    return xml_text


def _prof_29n_static_doctor_replacements(context: dict[str, str]) -> list[str]:
    return [
        _first_non_empty(context.get("Prof29ChairmanDoctor"), context.get("Doctor")),
        _first_non_empty(context.get("Prof29PathologistDoctor"), context.get("Prof29ChairmanDoctor"), context.get("Doctor")),
        context.get("Prof29PsychiatristNarcologistDoctor", ""),
        context.get("Prof29PsychiatristDoctor", ""),
    ]


def _replace_prof_29n_static_doctor_names(xml_text: str, context: dict[str, str]) -> str:
    if "Председатель врачебной комиссии" not in xml_text or "Психиат" not in xml_text:
        return xml_text

    for replacement in _prof_29n_static_doctor_replacements(context):
        replacement = str(replacement or "").strip()
        if not replacement:
            continue
        xml_text = re.sub(
            r"Сибирцев\s+В\.А",
            escape_xml_text(replacement),
            xml_text,
            count=1,
        )
    return xml_text


def _normalize_token_key(value: str) -> str:
    return re.sub(r"\s+", "", value.replace("|", ""))


def _token_key_pattern(key: str) -> str:
    return r"\s*\.\s*".join(re.escape(part) for part in key.split("."))


def _context_token_variants(context: dict[str, str]) -> list[tuple[str, str]]:
    variants: list[tuple[str, str]] = []
    seen: set[str] = set()
    for key, value in context.items():
        normalized = _normalize_token_key(key)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        variants.append((normalized, value))
    return variants


def _normalized_context_lookup(context: dict[str, str]) -> dict[str, str]:
    return {normalized: value for normalized, value in _context_token_variants(context)}


def _has_context_bookmarks(xml_text: str, context: dict[str, str]) -> bool:
    bookmark_names = re.findall(r'w:bookmarkStart\b[^>]*\bw:name="([^"]+)"', xml_text)
    if not bookmark_names:
        return False
    context_keys = set(context.keys())
    return any(name and not name.startswith("_") and name in context_keys for name in bookmark_names)


def _replace_text_tokens(xml_text: str, context: dict[str, str]) -> str:
    for key, value in _context_token_variants(context):
        key_pattern = _token_key_pattern(key)
        patterns = [
            rf"\[\s*\|\s*{key_pattern}\s*\|\s*\]",
            rf"\[\s*{key_pattern}\s*\]",
        ]
        for pattern in patterns:
            xml_text = re.sub(pattern, value, xml_text)
    return xml_text


def _replace_chairman_082_static_country(
    template_path: Path,
    xml_text: str,
    context: dict[str, str],
) -> str:
    if template_path.name.casefold() not in CHAIRMAN_CERTIFICATE_082_TEMPLATE_FILES:
        return xml_text
    country = escape_xml_text(str(context.get("Country", "") or "").strip())
    return xml_text.replace(">Болгария<", f">{country}<")


def _append_bookmark_value(tree: ET.ElementTree, context: dict[str, str]) -> ET.ElementTree:
    root = tree.getroot()
    for bookmark in root.findall(".//w:bookmarkStart", NS):
        bookmark_name = bookmark.attrib.get(f"{{{W_NS}}}name")
        if not bookmark_name or bookmark_name.startswith("_"):
            continue
        if bookmark_name not in context:
            continue

        parent = _find_parent(root, bookmark)
        if parent is None:
            continue

        bookmark_id = bookmark.attrib.get(f"{{{W_NS}}}id")
        value = context.get(bookmark_name, "")
        if value is None:
            value = ""

        # Remove existing text nodes until the matching bookmark end.
        removing = False
        to_remove: list[ET.Element] = []
        for child in list(parent):
            if child is bookmark:
                removing = True
                continue
            if removing and child.tag == f"{{{W_NS}}}bookmarkEnd" and child.attrib.get(f"{{{W_NS}}}id") == bookmark_id:
                break
            if removing and child.tag == f"{{{W_NS}}}r":
                to_remove.append(child)

        for node in to_remove:
            parent.remove(node)

        if value:
            run = ET.Element(f"{{{W_NS}}}r")
            text = ET.SubElement(run, f"{{{W_NS}}}t")
            text.text = value
            insert_index = list(parent).index(bookmark) + 1
            parent.insert(insert_index, run)
    return tree


def _replace_split_token_nodes(tree: ET.ElementTree, context: dict[str, str]) -> ET.ElementTree:
    root = tree.getroot()
    text_nodes = root.findall(".//w:t", NS)
    normalized_context = _normalized_context_lookup(context)
    index = 0

    while index < len(text_nodes):
        current_text = (text_nodes[index].text or "").strip()
        if current_text != "[":
            index += 1
            continue

        end_index = index + 1
        token_parts: list[str] = []
        found_end = False
        while end_index < len(text_nodes):
            part = (text_nodes[end_index].text or "").strip()
            if part == "]":
                found_end = True
                break
            token_parts.append(part)
            end_index += 1

        if not found_end:
            index += 1
            continue

        raw_token = "".join(token_parts)
        normalized_token = _normalize_token_key(raw_token)
        if normalized_token in normalized_context:
            text_nodes[index].text = normalized_context[normalized_token]
            for clear_index in range(index + 1, end_index + 1):
                text_nodes[clear_index].text = ""

        index = end_index + 1

    return tree


def _replace_text_node_tokens(tree: ET.ElementTree, context: dict[str, str]) -> ET.ElementTree:
    variants = _context_token_variants(context)
    for text_node in tree.getroot().findall(".//w:t", NS):
        text = text_node.text or ""
        if "[" not in text:
            continue
        for key, value in variants:
            key_pattern = _token_key_pattern(key)
            text = re.sub(rf"\[\s*\|\s*{key_pattern}\s*\|\s*\]", value, text)
            text = re.sub(rf"\[\s*{key_pattern}\s*\]", value, text)
        text_node.text = text
    return tree


def _replace_paragraph_tokens(tree: ET.ElementTree, context: dict[str, str]) -> ET.ElementTree:
    variants = _context_token_variants(context)
    for paragraph in tree.getroot().findall(".//w:p", NS):
        text_nodes = paragraph.findall(".//w:t", NS)
        if not text_nodes:
            continue
        text = "".join(node.text or "" for node in text_nodes)
        if "[" not in text:
            continue
        replaced = text
        for key, value in variants:
            key_pattern = _token_key_pattern(key)
            replaced = re.sub(rf"\[\s*\|\s*{key_pattern}\s*\|\s*\]", value, replaced)
            replaced = re.sub(rf"\[\s*{key_pattern}\s*\]", value, replaced)
        if replaced == text:
            continue
        text_nodes[0].text = replaced
        for node in text_nodes[1:]:
            node.text = ""
    return tree


def _find_parent(root: ET.Element, node: ET.Element) -> ET.Element | None:
    for parent in root.iter():
        for child in list(parent):
            if child is node:
                return parent
    return None


def _set_cell_text(cell: ET.Element, value: str) -> None:
    text_nodes = cell.findall(".//w:t", NS)
    if text_nodes:
        text_nodes[0].text = value
        for node in text_nodes[1:]:
            node.text = ""
        return

    paragraph = cell.find("w:p", NS)
    if paragraph is None:
        paragraph = ET.SubElement(cell, f"{{{W_NS}}}p")
    run = paragraph.find("w:r", NS)
    if run is None:
        run = ET.SubElement(paragraph, f"{{{W_NS}}}r")
    text = ET.SubElement(run, f"{{{W_NS}}}t")
    text.text = value


def _expand_service_rows(tree: ET.ElementTree, service_rows: list[dict[str, str]]) -> ET.ElementTree:
    root = tree.getroot()
    token = "qdfOrderServices_Ordinal_Service_Quantity_ServiceDate"
    for table in root.findall(".//w:tbl", NS):
        rows = table.findall("w:tr", NS)
        for row in rows:
            if len(row.findall("w:tc", NS)) < 4:
                continue
            row_text = "".join(text.text or "" for text in row.findall(".//w:t", NS))
            if token not in row_text:
                continue

            row_index = list(table).index(row)
            table.remove(row)
            if not service_rows:
                return tree
            for offset, item in enumerate(service_rows):
                new_row = ET.fromstring(ET.tostring(row, encoding="utf-8"))
                cells = new_row.findall("w:tc", NS)
                values = [
                    item.get("ordinal", ""),
                    item.get("service", ""),
                    item.get("quantity", "1"),
                    item.get("date", ""),
                ]
                for cell, value in zip(cells, values):
                    _set_cell_text(cell, value)
                table.insert(row_index + offset, new_row)
            return tree

    return tree


def _document_namespace_declarations(xml_text: str) -> list[tuple[str, str]]:
    document_start = xml_text.find("<w:document")
    if document_start < 0:
        return []
    document_start_end = xml_text.find(">", document_start)
    if document_start_end < 0:
        return []
    root_tag = xml_text[document_start:document_start_end]
    declarations = re.findall(
        r'\s+xmlns(?::([A-Za-z_][\w.-]*))?="([^"]+)"',
        root_tag,
    )
    for prefix, namespace_uri in declarations:
        if prefix == "xml":
            continue
        try:
            ET.register_namespace(prefix or "", namespace_uri)
        except ValueError:
            continue
    return declarations


def _restore_document_namespace_declarations(
    serialized_xml: bytes,
    declarations: list[tuple[str, str]],
) -> bytes:
    if not declarations:
        return serialized_xml

    xml_text = serialized_xml.decode("utf-8")
    document_start = xml_text.find("<w:document")
    if document_start < 0:
        return serialized_xml
    document_start_end = xml_text.find(">", document_start)
    if document_start_end < 0:
        return serialized_xml

    root_tag = xml_text[document_start:document_start_end]
    missing_attributes: list[str] = []
    for prefix, namespace_uri in declarations:
        attribute_name = f"xmlns:{prefix}" if prefix else "xmlns"
        if re.search(rf"\s+{re.escape(attribute_name)}=", root_tag):
            continue
        missing_attributes.append(f'{attribute_name}="{namespace_uri}"')
    if not missing_attributes:
        return serialized_xml

    insertion = " " + " ".join(missing_attributes)
    xml_text = xml_text[:document_start_end] + insertion + xml_text[document_start_end:]
    return xml_text.encode("utf-8")


def _generate_docx(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    service_rows: list[dict[str, str]] | None = None,
    cleanup_xml: bool = False,
) -> None:
    with zipfile.ZipFile(template_path, "r") as source_zip:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                file_bytes = source_zip.read(item.filename)
                if item.filename == "word/document.xml":
                    xml_text = file_bytes.decode("utf-8")
                    namespace_declarations = _document_namespace_declarations(xml_text)
                    xml_text = _replace_text_tokens(xml_text, context)
                    xml_text = _replace_chairman_082_static_country(template_path, xml_text, context)
                    xml_text = _replace_prof_29n_static_doctor_names(xml_text, context)
                    if cleanup_xml:
                        xml_text = _cleanup_contract_xml(xml_text)
                    needs_tree_pass = (
                        "[" in xml_text
                        or _has_context_bookmarks(xml_text, context)
                        or (
                            bool(service_rows)
                            and "qdfOrderServices_Ordinal_Service_Quantity_ServiceDate" in xml_text
                        )
                    )
                    if needs_tree_pass:
                        try:
                            tree = ET.ElementTree(ET.fromstring(xml_text))
                            tree = _replace_paragraph_tokens(tree, context)
                            tree = _replace_text_node_tokens(tree, context)
                            tree = _replace_split_token_nodes(tree, context)
                            tree = _append_bookmark_value(tree, context)
                            tree = _expand_service_rows(tree, service_rows or [])
                            file_bytes = _restore_document_namespace_declarations(
                                ET.tostring(tree.getroot(), encoding="utf-8", xml_declaration=True),
                                namespace_declarations,
                            )
                        except ParseError:
                            # Some client templates contain non-standard Word XML fragments.
                            # In that case we still keep token replacement instead of failing generation.
                            file_bytes = xml_text.encode("utf-8")
                    else:
                        file_bytes = xml_text.encode("utf-8")
                target_zip.writestr(item, file_bytes)


def _generate_xml(template_path: Path, output_path: Path, context: dict[str, str]) -> None:
    xml_text = template_path.read_text(encoding="utf-8", errors="ignore")
    xml_text = _replace_text_tokens(xml_text, context)
    for key, value in context.items():
        xml_text = xml_text.replace(f"{{{{{key}}}}}", value)
    output_path.write_text(xml_text, encoding="utf-8")


def _miac_xml_kind(template: DocumentTemplate) -> str | None:
    file_name = str(template.file_name or Path(template.file_path or "").name)
    if file_name.casefold() == MIAC_DRIVER_TEMPLATE.casefold():
        return "driver"
    if file_name.casefold() == MIAC_GUARD_TEMPLATE.casefold():
        return "guard"
    if file_name.casefold() == MIAC_GIMS_TEMPLATE.casefold():
        return "gims"
    return None


def _miac_text(value: object) -> str:
    return str(value or "").strip()


def _miac_add(parent: ET.Element, tag: str, value: object = "") -> ET.Element:
    node = ET.SubElement(parent, tag)
    node.text = _miac_text(value)
    return node


def _miac_mb_add(parent: ET.Element, tag: str, value: object = "") -> ET.Element:
    return _miac_add(parent, f"{{{MIAC_NS}}}{tag}", value)


def _miac_clean_locality(value: str) -> str:
    return re.sub(
        r"^\s*(?:г\.?|гор\.?|город|п\.?|пос\.?|поселок|посёлок|гп\.?|село|деревня)\s*",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()


def _miac_address_parts(address: str) -> dict[str, str]:
    parsed = _split_address(address)
    result = {
        "place": parsed.get("subject", ""),
        "area": parsed.get("district", ""),
        "city": _miac_clean_locality(parsed.get("city", "")),
        "town": "",
        "street": parsed.get("street", ""),
        "house": parsed.get("house", ""),
        "building": parsed.get("body", ""),
        "flat": parsed.get("apartment", ""),
    }

    town_pattern = re.compile(
        r"^\s*(?:п\.?|пос\.?|поселок|посёлок|гп\.?|село|деревня)\s+",
        re.IGNORECASE,
    )
    for part in (item.strip() for item in re.split(r",|\n", address or "")):
        if town_pattern.search(part):
            result["town"] = town_pattern.sub("", part).strip()
            if _miac_clean_locality(parsed.get("city", "")).casefold() == result["town"].casefold():
                result["city"] = ""
            break

    normalized = (address or "").casefold().replace("ё", "е")
    federal_city = ""
    if "санкт-петербург" in normalized or re.search(r"\bспб\b", normalized):
        federal_city = "Санкт-Петербург"
    elif "севастополь" in normalized:
        federal_city = "Севастополь"
    elif re.search(r"\bмосква\b", normalized):
        federal_city = "Москва"
    if federal_city:
        result["place"] = federal_city
        result["city"] = federal_city
    return result


def _miac_completed_chairman(exams: list[DoctorExam]) -> DoctorExam | None:
    return next(
        (
            exam
            for exam in exams
            if str(exam.doctor_role_id or "").strip().casefold() == "chairman"
            and bool(exam.is_completed)
            and exam.deleted_at is None
        ),
        None,
    )


def _miac_chairman_values(exams: list[DoctorExam]) -> tuple[DoctorExam | None, str, date | None, str]:
    chairman = _miac_completed_chairman(exams)
    if chairman is None:
        return None, "", None, ""
    fields = chairman.fields_json or {}
    conclusion = _first_non_empty(
        _exam_field(fields, "conclusionText", "issuedConclusion", "conclusion", "diagnosis"),
        chairman.result_text,
        chairman.diagnosis,
    )
    conclusion_date = None
    if chairman.completed_at:
        completed_at = chairman.completed_at
        if completed_at.tzinfo is not None:
            try:
                completed_at = completed_at.astimezone(ZoneInfo(settings.xml_exports_timezone))
            except (ZoneInfoNotFoundError, ValueError):
                pass
        conclusion_date = completed_at.date()
    return chairman, conclusion, conclusion_date, _miac_text(chairman.doctor_name)


def _validate_miac_values(
    *,
    client: Client,
    blank_form: BlankForm | None,
    chairman: DoctorExam | None,
    conclusion: str,
    conclusion_date: date | None,
    doctor_name: str,
) -> None:
    missing: list[str] = []
    if blank_form is None or not _miac_text(blank_form.full_number):
        missing.append("номер выданного бланка")
    if not _miac_text(client.last_name):
        missing.append("фамилия пациента")
    if not _miac_text(client.first_name):
        missing.append("имя пациента")
    if client.birth_date is None:
        missing.append("дата рождения пациента")
    if chairman is None:
        missing.append("завершённый осмотр председателя комиссии")
    if not conclusion:
        missing.append("заключение председателя комиссии")
    if conclusion_date is None:
        missing.append("дата заключения председателя комиссии")
    if not doctor_name:
        missing.append("ФИО председателя комиссии")
    if missing:
        raise ValueError("Невозможно сформировать XML МИАЦ. Заполните: " + ", ".join(missing))


def _miac_exam_result(exam: DoctorExam | None) -> str:
    if exam is None or not exam.is_completed or exam.deleted_at is not None:
        return ""
    data = _build_exam_export(exam)
    return _first_non_empty(data.get("diagnosis"), data.get("objective"), data.get("title"))


def _build_miac_guard_xml(
    client: Client,
    blank_form: BlankForm,
    exams: list[DoctorExam],
) -> ET.ElementTree:
    chairman, conclusion, conclusion_date, doctor_name = _miac_chairman_values(exams)
    _validate_miac_values(
        client=client,
        blank_form=blank_form,
        chairman=chairman,
        conclusion=conclusion,
        conclusion_date=conclusion_date,
        doctor_name=doctor_name,
    )
    address = _miac_address_parts(_miac_text(client.registration_text) or _miac_text(client.address_text))

    root = ET.Element("BlankSecurity")
    request = ET.SubElement(root, "Request")
    blank_info = ET.SubElement(request, "blankInfo")
    _miac_add(blank_info, "id", blank_form.full_number)
    duplicate = ET.SubElement(blank_info, "duplicate")
    _miac_add(duplicate, "duplicateId")
    _miac_add(duplicate, "isDuplicated", "false")
    _miac_add(blank_info, "isSpoiled", "false")

    security_blank = ET.SubElement(root, "SecurityBlank")
    user_info = ET.SubElement(security_blank, "userInfo")
    _miac_add(user_info, "surname", client.last_name)
    _miac_add(user_info, "name", client.first_name)
    _miac_add(user_info, "patronymic", client.middle_name)
    _miac_add(user_info, "birthday", client.birth_date.strftime("%d.%m.%Y"))
    address_node = ET.SubElement(security_blank, "address")
    for tag in ("place", "area", "city", "town", "street", "house", "building", "flat"):
        _miac_add(address_node, tag, address[tag])
    medical = ET.SubElement(security_blank, "medConclusion")
    _miac_add(medical, "conclusion", conclusion)
    _miac_add(medical, "dateConclusion", conclusion_date.strftime("%d.%m.%Y"))
    _miac_add(medical, "fioDoctor", doctor_name)
    return ET.ElementTree(root)


def _build_miac_driver_xml(
    client: Client,
    blank_form: BlankForm,
    exams: list[DoctorExam],
) -> ET.ElementTree:
    chairman, conclusion, conclusion_date, doctor_name = _miac_chairman_values(exams)
    _validate_miac_values(
        client=client,
        blank_form=blank_form,
        chairman=chairman,
        conclusion=conclusion,
        conclusion_date=conclusion_date,
        doctor_name=doctor_name,
    )
    fields = chairman.fields_json or {}
    actual_address = _miac_text(client.address_text)
    address = _miac_address_parts(actual_address or _miac_text(client.registration_text))
    address_type = "1" if actual_address else "0"

    root = ET.Element(f"{{{SOAP_NS}}}Envelope", {"xmlns:xd": XMLDSIG_NS})
    ET.SubElement(root, f"{{{SOAP_NS}}}Header")
    body = ET.SubElement(root, f"{{{SOAP_NS}}}Body")
    request = ET.SubElement(body, f"{{{MIAC_NS}}}fillGibddBlankV4Request")
    blank_info = _miac_mb_add(request, "blankInfo")
    _miac_mb_add(blank_info, "id", blank_form.full_number)
    _miac_mb_add(blank_info, "isSpoiled", "false")
    duplicate = _miac_mb_add(blank_info, "duplicate")
    _miac_mb_add(duplicate, "isDuplicated", "false")
    _miac_mb_add(duplicate, "duplicateId")

    client_info = _miac_mb_add(request, "clientInfo")
    _miac_mb_add(client_info, "surname", client.last_name)
    _miac_mb_add(client_info, "name", client.first_name)
    _miac_mb_add(client_info, "patronymic", client.middle_name)
    _miac_mb_add(client_info, "birthday", client.birth_date.isoformat())
    address_node = _miac_mb_add(client_info, "address")
    _miac_mb_add(address_node, "type", address_type)
    for tag in ("place", "area", "city", "town", "street", "house", "building", "flat"):
        _miac_mb_add(address_node, tag, address[tag])

    exams_by_role = _exam_map(exams)
    conclusion_node = _miac_mb_add(request, "conclusion")
    inspection = _miac_mb_add(conclusion_node, "inspectionResult")
    inspection_roles = (
        ("therapist", "therapist"),
        ("ophthalmologist", "ophthalmologist"),
        ("psychiatrist", "psychiatrist"),
        ("narcologist", "psychiatrist-narcologist"),
        ("neurologist", "neurologist"),
        ("otorhinolaryngologist", "otolaryngologist"),
    )
    for tag, role in inspection_roles:
        _miac_mb_add(inspection, tag, _miac_exam_result(exams_by_role.get(role)))
    _miac_mb_add(
        inspection,
        "instrumentalResearch",
        _exam_field(fields, "instrumentalResearch", "ekgConclusion", "ekg"),
    )
    _miac_mb_add(
        inspection,
        "laboratoryTest",
        _exam_field(fields, "laboratoryTest", "laboratoryStudy", "laboratoryTests"),
    )

    indication_keys = (
        "indicationManual",
        "indicationAutomatic",
        "indicationAcoustic",
        "indicationGlasses",
        "indicationHearingAid",
    )
    restriction_keys = ("restrictionAM", "restrictionBBE", "restrictionCCE")
    decision = _miac_text(fields.get("conclusion") or conclusion).casefold().replace("ё", "е")
    medical = _miac_mb_add(conclusion_node, "medConclusion")
    _miac_mb_add(
        medical,
        "contraindication",
        _bool_text(decision.startswith("не годен") or decision.startswith("негоден")),
    )
    _miac_mb_add(medical, "indication", _bool_text(any(_truthy_driver_value(fields.get(key)) for key in indication_keys)))
    _miac_mb_add(medical, "restriction", _bool_text(any(_truthy_driver_value(fields.get(key)) for key in restriction_keys)))
    _miac_mb_add(medical, "returnLicence", "false")
    _miac_mb_add(medical, "dateConclusion", conclusion_date.isoformat())
    _miac_mb_add(medical, "fioDoctor", doctor_name)

    category_node = _miac_mb_add(request, "category")
    categories = _miac_mb_add(category_node, "category")
    category_fields = (
        ("categoryA", "categoryA"),
        ("categoryB", "categoryB"),
        ("categoryC", "categoryC"),
        ("categoryD", "categoryD"),
        ("categoryBE", "categoryBE"),
        ("categoryCE", "categoryCE"),
        ("categoryDE", "categoryDE"),
        ("categoryTm", "categoryTram"),
        ("categoryTb", "categoryTrolleybus"),
        ("categoryM", "categoryM"),
        ("cat_tractor", "categoryTractor"),
    )
    for tag, field_key in category_fields:
        _miac_mb_add(categories, tag, _bool_text(_truthy_driver_value(fields.get(field_key))))
    _miac_mb_add(categories, "cat_vehicle", "false")
    _miac_mb_add(categories, "cat_ship", _bool_text(_truthy_driver_value(fields.get("categoryBoat"))))

    subcategories = _miac_mb_add(category_node, "subCategory")
    for tag, field_key in (
        ("subCategoryA1", "categoryA1"),
        ("subCategoryB1", "categoryB1"),
        ("subCategoryC1", "categoryC1"),
        ("subCategoryD1", "categoryD1"),
        ("subCategoryC1E", "categoryC1E"),
        ("subCategoryD1E", "categoryD1E"),
    ):
        _miac_mb_add(subcategories, tag, _bool_text(_truthy_driver_value(fields.get(field_key))))

    restrictions = _miac_mb_add(request, "restrictions")
    for tag, field_key in (("catAM", "restrictionAM"), ("catBBE", "restrictionBBE"), ("catCCED", "restrictionCCE")):
        _miac_mb_add(restrictions, tag, _bool_text(_truthy_driver_value(fields.get(field_key))))
    indications = _miac_mb_add(request, "indications")
    for tag, field_key in (
        ("manual", "indicationManual"),
        ("automatic", "indicationAutomatic"),
        ("parktronic", "indicationAcoustic"),
        ("correctVision", "indicationGlasses"),
        ("lossHearing", "indicationHearingAid"),
    ):
        _miac_mb_add(indications, tag, _bool_text(_truthy_driver_value(fields.get(field_key))))
    return ET.ElementTree(root)


def _build_miac_gims_xml(
    client: Client,
    blank_form: BlankForm,
    exams: list[DoctorExam],
) -> ET.ElementTree:
    chairman, conclusion, conclusion_date, doctor_name = _miac_chairman_values(exams)
    _validate_miac_values(
        client=client,
        blank_form=blank_form,
        chairman=chairman,
        conclusion=conclusion,
        conclusion_date=conclusion_date,
        doctor_name=doctor_name,
    )
    snils = "".join(char for char in _miac_text(getattr(client, "snils", "")) if char.isdigit())
    if len(snils) != 11:
        raise ValueError("Невозможно сформировать XML МИАЦ. Заполните: корректный СНИЛС пациента (11 цифр)")

    fields = chairman.fields_json or {}
    actual_address = _miac_text(client.address_text)
    address = _miac_address_parts(actual_address or _miac_text(client.registration_text))
    address_type = "1" if actual_address else "0"
    decision = _miac_text(fields.get("conclusion") or conclusion).casefold().replace("ё", "е")
    contraindication = decision.startswith("не годен") or decision.startswith("негоден") or "противопоказан" in decision
    restriction_keys = (
        "restrictionOnManagement",
        "restrictionAM",
        "restrictionBBE",
        "restrictionCCE",
        "restrictionNoHands",
        "restrictionNoLegs",
    )
    reexamination_keys = ("reexaminationAfterBan", "returnLicence", "returnLicense")

    root = ET.Element(f"{{{MIAC_NS}}}fillShipBlankRequset")
    blank_info = _miac_mb_add(root, "blankInfo")
    _miac_mb_add(blank_info, "id", blank_form.full_number)
    duplicate = _miac_mb_add(blank_info, "duplicate")
    _miac_mb_add(duplicate, "isDuplicated", "false")
    _miac_mb_add(duplicate, "duplicateId")
    _miac_mb_add(blank_info, "isSpoiled", "false")

    client_info = _miac_mb_add(root, "clientInfo")
    _miac_mb_add(client_info, "surname", client.last_name)
    _miac_mb_add(client_info, "name", client.first_name)
    _miac_mb_add(client_info, "patronymic", client.middle_name)
    _miac_mb_add(client_info, "birthday", client.birth_date.strftime("%d.%m.%Y"))
    _miac_mb_add(client_info, "snils", snils)
    address_node = _miac_mb_add(client_info, "address")
    _miac_mb_add(address_node, "type", address_type)
    for tag in ("place", "area"):
        _miac_mb_add(address_node, tag, address[tag])
    _miac_mb_add(address_node, "town", address["town"] or address["city"])
    for tag in ("street", "house", "building", "flat"):
        _miac_mb_add(address_node, tag, address[tag])

    conclusion_node = _miac_mb_add(root, "conclusion")
    medical = _miac_mb_add(conclusion_node, "medConclusion")
    _miac_mb_add(medical, "contraindicationToManagement", _bool_text(contraindication))
    _miac_mb_add(
        medical,
        "restrictionOnManagement",
        _bool_text(any(_truthy_driver_value(fields.get(key)) for key in restriction_keys)),
    )
    _miac_mb_add(
        medical,
        "reexaminationAfterBan",
        _bool_text(any(_truthy_driver_value(fields.get(key)) for key in reexamination_keys)),
    )
    _miac_mb_add(medical, "dateConclusion", conclusion_date.strftime("%d.%m.%Y"))
    _miac_mb_add(medical, "fioDoctor", doctor_name)
    return ET.ElementTree(root)


def _generate_miac_xml(
    output_path: Path,
    *,
    kind: str,
    client: Client,
    blank_form: BlankForm,
    exams: list[DoctorExam],
) -> None:
    if kind == "driver":
        tree = _build_miac_driver_xml(client, blank_form, exams)
    elif kind == "gims":
        tree = _build_miac_gims_xml(client, blank_form, exams)
    else:
        tree = _build_miac_guard_xml(client, blank_form, exams)
    ET.indent(tree, space="   ")
    tree.write(output_path, encoding="utf-8", xml_declaration=True, short_empty_elements=True)


def _resolve_miac_issued_blank(
    db: Session,
    *,
    blank_type: str,
    client_id: int,
    encounter_id: int,
    center_id: int,
    blank_form_id: int | None,
) -> BlankForm:
    if blank_form_id is not None:
        form = db.get(BlankForm, blank_form_id)
    else:
        form = db.execute(
            select(BlankForm)
            .where(
                BlankForm.blank_type == blank_type,
                BlankForm.client_id == client_id,
                BlankForm.encounter_id == encounter_id,
                BlankForm.center_id == center_id,
                BlankForm.status == BLANK_STATUS_ISSUED,
            )
            .order_by(BlankForm.issued_at.desc(), BlankForm.id.desc())
            .limit(1)
        ).scalar_one_or_none()
    if form is None:
        raise ValueError("XML МИАЦ формируется только для уже выданного номерного бланка")
    if (
        form.status != BLANK_STATUS_ISSUED
        or form.blank_type != blank_type
        or form.client_id != client_id
        or form.encounter_id != encounter_id
        or form.center_id != center_id
    ):
        raise ValueError("Выбранный бланк не является выданным бланком этого пациента и обращения")
    return form


def _xml_export_date_folder() -> str:
    try:
        tzinfo = ZoneInfo(settings.xml_exports_timezone)
    except ZoneInfoNotFoundError:
        tzinfo = ZoneInfo("Europe/Moscow")
    return datetime.now(tzinfo).date().isoformat()


class _XlsxSheetAdapter:
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.name = worksheet.title

    @property
    def nrows(self) -> int:
        return self.worksheet.max_row

    @property
    def ncols(self) -> int:
        return self.worksheet.max_column

    @property
    def merged_cells(self) -> list[tuple[int, int, int, int]]:
        return [
            (cell_range.min_row - 1, cell_range.max_row, cell_range.min_col - 1, cell_range.max_col)
            for cell_range in self.worksheet.merged_cells.ranges
        ]

    def cell_value(self, row_index: int, col_index: int):
        value = self.worksheet.cell(row=row_index + 1, column=col_index + 1).value
        return "" if value is None else value

    def _writable_cell(self, row_index: int, col_index: int):
        row = row_index + 1
        column = col_index + 1
        cell = self.worksheet.cell(row=row, column=column)
        if not isinstance(cell, MergedCell):
            return cell
        for cell_range in self.worksheet.merged_cells.ranges:
            if cell.coordinate in cell_range:
                return self.worksheet.cell(row=cell_range.min_row, column=cell_range.min_col)
        return cell

    def write_cell(self, row_index: int, col_index: int, value: object) -> None:
        cell = self._writable_cell(row_index, col_index)
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    def write(self, row_index: int, col_index: int, value: object, style=None) -> None:
        self.write_cell(row_index, col_index, value)

    def write_merge(
        self,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        value: object,
        style=None,
    ) -> None:
        start = f"{get_column_letter(start_col + 1)}{start_row + 1}"
        end = f"{get_column_letter(end_col + 1)}{end_row + 1}"
        range_name = f"{start}:{end}"
        if not any(str(cell_range) == range_name for cell_range in self.worksheet.merged_cells.ranges):
            self.worksheet.merge_cells(range_name)
        self.write_cell(start_row, start_col, value)


class _XlsxBookAdapter:
    is_xlsx = True

    def __init__(self, workbook):
        self.workbook = workbook

    def sheet_names(self) -> list[str]:
        return list(self.workbook.sheetnames)

    def sheets(self) -> list[_XlsxSheetAdapter]:
        return [_XlsxSheetAdapter(sheet) for sheet in self.workbook.worksheets]

    def sheet_by_index(self, index: int) -> _XlsxSheetAdapter:
        return _XlsxSheetAdapter(self.workbook.worksheets[index])

    def get_sheet(self, index: int) -> _XlsxSheetAdapter:
        return self.sheet_by_index(index)

    def save(self, output_path: str | Path) -> None:
        self.workbook.save(output_path)

    def keep_only_sheets(self, target_sheet_names: tuple[str, ...]) -> None:
        kept_sheet = next((self.workbook[name] for name in target_sheet_names if name in self.workbook.sheetnames), None)
        if kept_sheet is None:
            raise ValueError(f"В шаблоне не найден лист для печати: {target_sheet_names[0]}")
        for sheet in list(self.workbook.worksheets):
            if sheet is not kept_sheet:
                self.workbook.remove(sheet)
        self.workbook.active = 0

    def iter_auto_markers(self) -> list[tuple[int, object, int, int, str]]:
        markers: list[tuple[int, object, int, int, str]] = []
        for sheet_index, worksheet in enumerate(self.workbook.worksheets):
            source_sheet = _XlsxSheetAdapter(worksheet)
            seen_merges: set[str] = set()
            for row in worksheet.iter_rows():
                for cell in row:
                    if not _xlsx_cell_is_yellow(cell):
                        continue
                    row_index = cell.row - 1
                    col_index = cell.column - 1
                    merge = next((cell_range for cell_range in worksheet.merged_cells.ranges if cell.coordinate in cell_range), None)
                    if merge is not None:
                        merge_key = str(merge)
                        if merge_key in seen_merges:
                            continue
                        seen_merges.add(merge_key)
                        row_index = merge.min_row - 1
                        col_index = merge.min_col - 1
                    label = _normalize_xls_auto_label(source_sheet.cell_value(row_index, col_index))
                    if "авто" in label:
                        markers.append((sheet_index, source_sheet, row_index, col_index, label))
        return markers


def _xlsx_cell_is_yellow(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    color = fill.fgColor
    if color is None:
        return False
    if color.type == "indexed":
        return color.indexed == 13
    rgb = str(color.rgb or "").upper()
    return rgb in {"FFFFFF00", "FFFF00"}


def _write_xls_cell(target_sheet, source_sheet, row_index: int, col_index: int, value: object, style=None) -> None:
    if hasattr(target_sheet, "write_cell"):
        target_sheet.write_cell(row_index, col_index, value)
        return

    existing_xf_idx = None
    existing_row = target_sheet._Worksheet__rows.get(row_index)
    if existing_row is not None:
        existing_cell = existing_row._Row__cells.get(col_index)
        if existing_cell is not None:
            existing_xf_idx = getattr(existing_cell, "xf_idx", None)

    if style is not None:
        target_sheet.write(row_index, col_index, value, style)
        return

    target_sheet.write(row_index, col_index, value)
    row = target_sheet._Worksheet__rows.get(row_index)
    if row is None:
        return
    cell = row._Row__cells.get(col_index)
    if cell is None:
        return
    if existing_xf_idx is not None:
        cell.xf_idx = existing_xf_idx
        return
    try:
        cell.xf_idx = source_sheet.cell_xf_index(row_index, col_index)
    except IndexError:
        # Some legacy sheets have blank trailing rows without style metadata.
        # In that case we keep the written value without cloning formatting.
        return


def _xls_cell_style(
    source_book,
    source_sheet,
    row_index: int,
    col_index: int,
    *,
    shrink_to_fit: bool | None = None,
    num_format_str: str | None = None,
):
    if getattr(source_book, "is_xlsx", False):
        return None

    xf = source_book.xf_list[source_sheet.cell_xf_index(row_index, col_index)]
    style = xlwt.XFStyle()

    if num_format_str is not None:
        style.num_format_str = num_format_str
    else:
        try:
            style.num_format_str = source_book.format_map[xf.format_key].format_str
        except KeyError:
            pass

    source_font = source_book.font_list[xf.font_index]
    font = xlwt.Font()
    font.height = source_font.height
    font.italic = bool(source_font.italic)
    font.struck_out = bool(source_font.struck_out)
    font.outline = bool(source_font.outline)
    font.shadow = bool(source_font.shadow)
    font.colour_index = source_font.colour_index
    font.bold = bool(source_font.bold)
    font._weight = source_font.weight
    font.escapement = source_font.escapement
    font.underline = source_font.underline_type
    font.family = source_font.family
    font.charset = source_font.character_set
    font.name = source_font.name
    style.font = font

    source_alignment = xf.alignment
    alignment = xlwt.Alignment()
    alignment.horz = source_alignment.hor_align
    alignment.vert = source_alignment.vert_align
    alignment.dire = source_alignment.text_direction
    alignment.rota = source_alignment.rotation
    alignment.wrap = source_alignment.text_wrapped
    should_shrink = source_alignment.shrink_to_fit if shrink_to_fit is None else shrink_to_fit
    alignment.shri = xlwt.Alignment.SHRINK_TO_FIT if should_shrink else 0
    alignment.inde = source_alignment.indent_level
    style.alignment = alignment

    source_border = xf.border
    borders = xlwt.Borders()
    borders.left = source_border.left_line_style
    borders.right = source_border.right_line_style
    borders.top = source_border.top_line_style
    borders.bottom = source_border.bottom_line_style
    borders.diag = source_border.diag_line_style
    borders.left_colour = source_border.left_colour_index
    borders.right_colour = source_border.right_colour_index
    borders.top_colour = source_border.top_colour_index
    borders.bottom_colour = source_border.bottom_colour_index
    borders.diag_colour = source_border.diag_colour_index
    borders.need_diag1 = source_border.diag_down
    borders.need_diag2 = source_border.diag_up
    style.borders = borders

    source_background = xf.background
    pattern = xlwt.Pattern()
    pattern.pattern = source_background.fill_pattern
    pattern.pattern_fore_colour = source_background.pattern_colour_index
    pattern.pattern_back_colour = source_background.background_colour_index
    style.pattern = pattern

    source_protection = xf.protection
    protection = xlwt.Protection()
    protection.cell_locked = source_protection.cell_locked
    protection.formula_hidden = source_protection.formula_hidden
    style.protection = protection

    return style


def _xls_shrink_to_fit_style(source_book, source_sheet, row_index: int, col_index: int):
    return _xls_cell_style(source_book, source_sheet, row_index, col_index, shrink_to_fit=True)


def _copy_xls_target_cell_style(target_sheet, target_row: int, target_col: int, source_row: int, source_col: int) -> None:
    if hasattr(target_sheet, "write_cell"):
        return

    target_row_obj = target_sheet._Worksheet__rows.get(target_row)
    if target_row_obj is None:
        return
    target_cell = target_row_obj._Row__cells.get(target_col)
    if target_cell is None:
        return

    source_row_obj = target_sheet._Worksheet__rows.get(source_row)
    if source_row_obj is None:
        return
    source_cell = source_row_obj._Row__cells.get(source_col)
    if source_cell is None:
        return
    source_xf_idx = getattr(source_cell, "xf_idx", None)
    if source_xf_idx is not None:
        target_cell.xf_idx = source_xf_idx


def _xls_excel_date(value: date | datetime | str | None) -> float | str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if not isinstance(value, date):
        return str(value)
    if value <= date(1900, 1, 1):
        return ""
    if value < date(1900, 3, 1):
        return value.strftime("%d.%m.%y")
    return xldate.xldate_from_date_tuple((value.year, value.month, value.day), 0)


def _prof_xls_display_date(value: date | datetime | str | None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%d.%m.%y")
    return str(value)


def _exam_field(fields: dict, *keys: str) -> str:
    lowered = {str(key).lower(): value for key, value in fields.items()}
    for key in keys:
        value = fields.get(key)
        if value in (None, ""):
            value = lowered.get(key.lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _build_exam_export(exam: DoctorExam | None) -> dict[str, object]:
    if exam is None:
        return {
            "date": "",
            "title": "",
            "complaints": "",
            "anamnesis": "",
            "objective": "",
            "diagnosis": "",
            "doctor": "",
        }

    fields = exam.fields_json or {}
    result_text = str(exam.result_text or "").strip()
    diagnosis = (
        str(exam.diagnosis or "").strip()
        or _exam_field(fields, "diagnosis", "diagnosisShort", "diagnosisText", "diagnoz", "conclusion")
        or result_text
    )
    objective = (
        _exam_field(
            fields,
            "objective",
            "objectiveData",
            "objectiveText",
            "status",
            "statusLocalis",
            "inspection",
            "exam",
            "result",
        )
        or result_text
    )
    return {
        "date": exam.completed_at.date() if exam.completed_at else "",
        "title": _exam_field(fields, "conclusionTitle", "title", "caption"),
        "complaints": _exam_field(fields, "complaints", "complaint", "complaintsText"),
        "anamnesis": _exam_field(fields, "anamnesis", "anamnesisText", "history", "anamnesisVitae"),
        "objective": objective,
        "diagnosis": diagnosis,
        "doctor": str(exam.doctor_name or "").strip(),
    }


def _prof_extract_exam_date(exam: DoctorExam, encounter: Encounter | None) -> object:
    if exam.completed_at:
        return exam.completed_at.date()
    if encounter is not None:
        return encounter.encounter_date
    return ""


def _prof_extract_exam_conclusion(exam: DoctorExam) -> str:
    fields = exam.fields_json or {}
    conclusion = _exam_field(fields, "conclusion", "result", "decision", "fit")
    if not conclusion:
        conclusion = str(exam.result_text or exam.diagnosis or "").strip()

    normalized = conclusion.strip().lower()
    if normalized == "годен":
        return "годен"
    if normalized == "не годен":
        return "не годен"
    return conclusion


def _prof_extract_doctor_name(client: Client | None, role_id: str, exam_data: dict[str, object]) -> str:
    client_field = PROF_EXTRACT_CLIENT_DOCTOR_FIELDS.get(role_id)
    if client is not None and client_field:
        client_value = str(getattr(client, client_field, "") or "").strip()
        if client_value:
            return client_value
    return str(exam_data.get("doctor") or "").strip()


def _exam_export_with_client_doctor(
    exam: DoctorExam | None,
    client: Client | None,
    role_id: str,
) -> dict[str, object]:
    data = _build_exam_export(exam)
    data["doctor"] = _prof_extract_doctor_name(client, role_id, data)
    return data


def _exam_conclusion_line_for_role(
    exam: DoctorExam | None,
    client: Client | None,
    role_id: str,
    fallback: str = "Противопоказания отсутствуют",
) -> str:
    data = _exam_export_with_client_doctor(exam, client, role_id)
    doctor = str(data.get("doctor") or "").strip()
    conclusion = _first_non_empty(data.get("diagnosis"), data.get("objective"), data.get("title"), fallback)
    return " ".join(part for part in [doctor, conclusion] if part).strip()


def _prof_extract_doctor_row_values(
    exams_by_role: dict[str, DoctorExam],
    encounter: Encounter | None,
    client: Client | None = None,
) -> list[tuple[int, str, object, str]]:
    rows: list[tuple[int, str, object, str]] = []
    row_indices = [row_index for _, _, row_index in PROF_EXTRACT_DOCTOR_ROWS]
    for role_id, specialty, row_index in PROF_EXTRACT_DOCTOR_ROWS:
        exam = exams_by_role.get(role_id)
        if exam is None or not exam.is_completed:
            continue

        exam_data = _build_exam_export(exam)
        doctor_name = _prof_extract_doctor_name(client, role_id, exam_data)
        doctor_line = " ".join(part for part in [specialty, doctor_name] if part).strip()
        target_row_index = row_indices[len(rows)]
        rows.append(
            (
                target_row_index,
                doctor_line,
                _prof_extract_exam_date(exam, encounter),
                _prof_extract_exam_conclusion(exam),
            )
        )
    return rows


def _prof_amb_exam_block_values(
    exams_by_role: dict[str, DoctorExam],
    encounter: Encounter | None,
    client: Client | None = None,
) -> list[tuple[str, dict[str, object]]]:
    blocks: list[tuple[str, dict[str, object]]] = []
    for role_id, specialty, _ in PROF_EXTRACT_DOCTOR_ROWS:
        exam = exams_by_role.get(role_id)
        if exam is None or not exam.is_completed:
            continue

        data = _build_exam_export(exam)
        data["doctor"] = _prof_extract_doctor_name(client, role_id, data)
        if not data.get("date"):
            data["date"] = _prof_extract_exam_date(exam, encounter)
        if not str(data.get("title") or "").strip():
            data["title"] = f"Врач {specialty.lower()}"
        blocks.append((role_id, data))
    return blocks


def _is_rural_address(*parts: str) -> bool:
    text = " ".join(part for part in parts if part).lower()
    return bool(re.search(r"\b(пос|пгт|село|дер|деревня|снт|рп|гп)\b", text))


def _split_policy(policy: str) -> tuple[str, str]:
    cleaned = re.sub(r"\s+", " ", str(policy or "").strip())
    if not cleaned:
        return "", ""
    parts = cleaned.split(" ", 1)
    if len(parts) == 1:
        digits = re.sub(r"\D", "", parts[0])
        if len(digits) > 10:
            return digits[:-10], digits[-10:]
        return "", parts[0]
    return parts[0], parts[1]


def _split_document(series: str, number: str) -> tuple[str, str]:
    merged = " ".join(part for part in [str(series or "").strip(), str(number or "").strip()] if part).strip()
    digits = re.sub(r"\D", "", merged)
    if len(digits) >= 10:
        return digits[:4], digits[4:10]
    return str(series or "").strip(), str(number or "").strip()


def _clear_xls_cells(target_sheet, source_sheet, coordinates: list[tuple[int, int]]) -> None:
    for row_index, col_index in coordinates:
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, "")


def _prof_amb_doctor_name_range(doctor_cell: tuple[int, int]) -> tuple[int, int, int]:
    row_index, col_index = doctor_cell
    end_col = 62 if col_index >= 32 else 31
    return row_index, col_index, end_col


def _fill_exam_block(
    target_sheet,
    source_sheet,
    data: dict[str, object],
    *,
    source_book=None,
    date_cell: tuple[int, int],
    title_cell: tuple[int, int],
    complaints_cell: tuple[int, int],
    anamnesis_cell: tuple[int, int],
    objective_cell: tuple[int, int],
    diagnosis_cell: tuple[int, int],
    doctor_cell: tuple[int, int],
) -> None:
    date_style = (
        _xls_cell_style(source_book, source_sheet, *date_cell, shrink_to_fit=False, num_format_str="@")
        if source_book is not None
        else None
    )
    _write_xls_cell(target_sheet, source_sheet, *date_cell, _prof_xls_display_date(data.get("date")), date_style)
    _write_xls_cell(target_sheet, source_sheet, *title_cell, str(data.get("title") or ""))
    _write_xls_cell(target_sheet, source_sheet, *complaints_cell, str(data.get("complaints") or ""))
    _write_xls_cell(target_sheet, source_sheet, *anamnesis_cell, str(data.get("anamnesis") or ""))
    _write_xls_cell(target_sheet, source_sheet, *objective_cell, str(data.get("objective") or ""))
    _write_xls_cell(target_sheet, source_sheet, *diagnosis_cell, str(data.get("diagnosis") or ""))
    doctor_value = str(data.get("doctor") or "").strip()
    doctor_row, doctor_start_col, doctor_end_col = _prof_amb_doctor_name_range(doctor_cell)
    doctor_style = (
        _xls_cell_style(source_book, source_sheet, *doctor_cell, shrink_to_fit=False)
        if source_book is not None
        else None
    )
    if doctor_value and source_book is not None and doctor_end_col > doctor_start_col:
        target_sheet.write_merge(doctor_row, doctor_row, doctor_start_col, doctor_end_col, doctor_value, doctor_style)
    elif doctor_value:
        _write_xls_cell(target_sheet, source_sheet, *doctor_cell, doctor_value, doctor_style)
    else:
        _clear_xls_cells(
            target_sheet,
            source_sheet,
            [(doctor_row, col_index) for col_index in range(doctor_start_col, doctor_end_col + 1)],
        )


def _clear_prof_amb_exam_block(target_sheet, source_sheet, block: dict[str, tuple[int, int]]) -> None:
    date_cell = block["date_cell"]
    title_cell = block["title_cell"]
    doctor_cell = block["doctor_cell"]
    label_col = max(0, title_cell[1] - 9)
    end_col = min(source_sheet.ncols, label_col + 31)
    coordinates: set[tuple[int, int]] = set(block.values())
    doctor_row, doctor_start_col, doctor_end_col = _prof_amb_doctor_name_range(doctor_cell)
    coordinates.update((doctor_row, col_index) for col_index in range(doctor_start_col, doctor_end_col + 1))
    for row_index in range(date_cell[0], doctor_cell[0] + 1):
        for col_index in range(label_col, end_col):
            if source_sheet.cell_value(row_index, col_index) not in ("", None):
                coordinates.add((row_index, col_index))
    _clear_xls_cells(target_sheet, source_sheet, sorted(coordinates))


def _write_xls_pairs(target_sheet, source_sheet, pairs: list[tuple[tuple[int, int], object]]) -> None:
    for (row_index, col_index), value in pairs:
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, value)


def _normalize_xls_auto_label(value: object) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", " ", text).strip()


def _iter_xls_auto_markers(source_book) -> list[tuple[int, object, int, int, str]]:
    if hasattr(source_book, "iter_auto_markers"):
        return source_book.iter_auto_markers()

    markers: list[tuple[int, object, int, int, str]] = []
    for sheet_index, source_sheet in enumerate(source_book.sheets()):
        seen_merges: set[tuple[int, int, int, int]] = set()
        for row_index in range(source_sheet.nrows):
            for col_index in range(source_sheet.ncols):
                try:
                    xf_index = source_sheet.cell_xf_index(row_index, col_index)
                    bg_index = source_book.xf_list[xf_index].background.pattern_colour_index
                except IndexError:
                    continue
                if bg_index != 13:
                    continue

                merge = next(
                    (
                        item
                        for item in source_sheet.merged_cells
                        if item[0] <= row_index < item[1] and item[2] <= col_index < item[3]
                    ),
                    None,
                )
                if merge is not None:
                    if merge in seen_merges:
                        continue
                    seen_merges.add(merge)
                    row_index, col_index = merge[0], merge[2]

                label = _normalize_xls_auto_label(source_sheet.cell_value(row_index, col_index))
                if "авто" in label:
                    markers.append((sheet_index, source_sheet, row_index, col_index, label))
    return markers


def _xls_auto_marker_values(
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> list[tuple[tuple[str, ...], object]]:
    issue_date = encounter.encounter_date if encounter else date.today()
    therapist = _exam_export_with_client_doctor(exams_by_role.get("therapist"), client, "therapist")
    chairman = _build_exam_export(exams_by_role.get("chairman"))
    return [
        (("терапевт",), _exam_conclusion_line_for_role(exams_by_role.get("therapist"), client, "therapist")),
        (("офтальмолог", "окулист"), _exam_conclusion_line_for_role(exams_by_role.get("ophthalmologist"), client, "ophthalmologist")),
        (("невролог",), _exam_conclusion_line_for_role(exams_by_role.get("neurologist"), client, "neurologist")),
        (("лор", "отоларинголог"), _exam_conclusion_line_for_role(exams_by_role.get("otolaryngologist"), client, "otolaryngologist")),
        (("хирург",), _exam_conclusion_line_for_role(exams_by_role.get("surgeon"), client, "surgeon")),
        (("психиатр нарколог", "психиатр-нарколог", "нарколог"), _exam_conclusion_line_for_role(exams_by_role.get("psychiatrist-narcologist"), client, "psychiatrist-narcologist")),
        (("психиатр",), _exam_conclusion_line_for_role(exams_by_role.get("psychiatrist"), client, "psychiatrist")),
        (("дермат",), _exam_conclusion_line_for_role(exams_by_role.get("dermatologist"), client, "dermatologist")),
        (("гинеколог",), _exam_conclusion_line_for_role(exams_by_role.get("gynecologist"), client, "gynecologist")),
        (("председатель", "глав врач", "главный врач", "подписант"), _first_non_empty(chairman.get("doctor"), therapist.get("doctor"), context.get("Doctor"))),
        (("врач",), _first_non_empty(therapist.get("doctor"), context.get("Doctor"))),
        (("фио", "пациент"), context.get("ClientCalc", "")),
        (("дата рождения", "др"), _xls_excel_date(client.birth_date)),
        (("возраст",), _age_at_date(client.birth_date, issue_date)),
        (("пол",), context.get("SexCalc", "")),
        (("адрес",), context.get("AddressCalc", "")),
        (("телефон",), context.get("Phone", "")),
        (("снилс",), context.get("SNILS", "")),
        (("паспорт серия", "серия"), context.get("DocumentSeries", "")),
        (("паспорт номер", "номер паспорта"), context.get("DocumentNumber", "")),
        (("кем выдан",), context.get("WhoGive", "")),
        (("дата выдачи",), context.get("DocumentDate", "")),
        (("организация", "место работы", "работа"), context.get("CompanyName", "")),
        (("должность", "профессия"), _first_non_empty(context.get("Post"), context.get("PositionApplied"))),
        (("услуга", "услуги"), context.get("Services", "")),
        (("номер бланка", "бланк"), context.get("BlankFullNumber") or context.get("BlankNumber", "")),
        (("номер",), context.get("ReferenceNumber", "")),
        (("дата",), _xls_excel_date(issue_date)),
        (("заключение", "итог"), context.get("Conclusion", "")),
    ]


def _apply_xls_auto_markers(
    source_book,
    target_book,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    values = _xls_auto_marker_values(context, client, encounter, exams_by_role)
    for sheet_index, source_sheet, row_index, col_index, label in _iter_xls_auto_markers(source_book):
        value = None
        for aliases, candidate in values:
            if any(alias in label for alias in aliases):
                value = candidate
                break
        if value is None:
            continue
        target_sheet = target_book.get_sheet(sheet_index)
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, value)


def _exam_map(exams: list[DoctorExam]) -> dict[str, DoctorExam]:
    result: dict[str, DoctorExam] = {}
    for exam in exams:
        role_key = str(exam.doctor_role_id or "").strip().lower()
        result.setdefault(role_key, exam)
    return result


def _first_non_empty(*values: object) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _exam_conclusion_line(exam: DoctorExam | None, fallback: str = "Противопоказания отсутствуют") -> str:
    data = _build_exam_export(exam)
    doctor = str(data.get("doctor") or "").strip()
    conclusion = _first_non_empty(data.get("diagnosis"), data.get("objective"), data.get("title"), fallback)
    return " ".join(part for part in [doctor, conclusion] if part).strip()


def _age_at_date(birth_date: date | None, on_date: date | None) -> str:
    if birth_date is None:
        return ""
    check_date = on_date or date.today()
    years = check_date.year - birth_date.year - ((check_date.month, check_date.day) < (birth_date.month, birth_date.day))
    return str(years)


def _sheet_pair(source_book, target_book, sheet_name: str):
    if sheet_name not in source_book.sheet_names():
        return None, None, None
    index = source_book.sheet_names().index(sheet_name)
    return source_book.sheet_by_index(index), target_book.get_sheet(index), index


def _sheet_pair_any(source_book, target_book, sheet_names: tuple[str, ...]):
    for sheet_name in sheet_names:
        source_sheet, target_sheet, index = _sheet_pair(source_book, target_book, sheet_name)
        if source_sheet is not None and target_sheet is not None:
            return source_sheet, target_sheet, index
    return None, None, None


def _fill_contract_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    runtime_values: dict[str, object],
) -> None:
    service_rows = runtime_values.get("service_rows", [])
    first_service = service_rows[0]["service"] if service_rows else context.get("Services", "")
    quantity = service_rows[0]["quantity"] if service_rows else "1"
    total_amount = str(encounter.total_amount or "") if encounter else ""
    doctor_name = context.get("UserName", "")
    full_name = context.get("ClientCalc", "")
    birth_date = context.get("BirthDateCalc", "")
    address = context.get("AddressCalc", "")
    passport_summary = (
        f"Паспорт РФ Серия:{context.get('DocumentSeries', '')} "
        f"Номер:{context.get('DocumentNumber', '')} "
        f"Кем выдан: {context.get('WhoGive', '')} - {context.get('DocumentDate', '')}"
    ).strip()
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((6, 2), f"Я, {full_name}, {birth_date} г. рождения,"),
            ((8, 6), address),
            ((8, 33), f"/ {doctor_name}" if doctor_name else ""),
            ((12, 34), context.get("ReferenceNumber", "")),
            ((14, 37), _xls_excel_date(encounter.encounter_date if encounter else date.today())),
            ((21, 1), f"{full_name} ( {context.get('Phone', '')} )"),
            ((24, 1), full_name),
            ((27, 16), _xls_excel_date(encounter.encounter_date if encounter else date.today())),
            ((31, 1), f"Я, {full_name}, {birth_date} г. рождения, зарегистрированная по адресу:"),
            ((32, 1), address),
            ((33, 1), passport_summary),
            ((33, 23), first_service),
            ((33, 37), quantity),
            ((33, 39), total_amount),
            ((37, 28), f"/ {full_name}" if full_name else ""),
            ((48, 1), f"Настоящее согласие дано мной {context.get('VisitDate', '')} и действует бессрочно."),
            ((55, 1), full_name),
            ((82, 23), total_amount),
            ((82, 33), f"/ {doctor_name}" if doctor_name else ""),
            ((86, 23), total_amount),
            ((86, 33), f"/ {doctor_name}" if doctor_name else ""),
            ((89, 33), f"Ф.И.О.: {full_name}" if full_name else ""),
            ((91, 33), f"Адрес места жительства: {address}" if address else ""),
            ((93, 35), client.document_type or "Паспорт РФ"),
            ((94, 35), context.get("DocumentSeries", "")),
            ((95, 35), context.get("DocumentNumber", "")),
            ((96, 33), f"Кем выдан: {context.get('WhoGive', '')}".strip()),
            ((99, 35), _xls_excel_date(client.document_issued_date)),
            ((100, 35), context.get("Phone", "")),
            ((106, 36), f"/ {full_name}" if full_name else ""),
        ],
    )


def _fill_086_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    issue_date = encounter.encounter_date if encounter else date.today()
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((10, 12), context.get("ReferenceNumber", "")),
            ((13, 5), context.get("ClientCalc", "")),
            ((14, 5), _xls_excel_date(client.birth_date)),
            ((15, 10), context.get("SubjectCalc", "")),
            ((16, 3), context.get("DistrictCalc", "")),
            ((17, 4), context.get("CityCalc", "")),
            ((17, 9), " ".join(
                part
                for part in [
                    context.get("StreetCalc", ""),
                    context.get("HouseNumberCalc", ""),
                    context.get("ApartmentNumberCalc", ""),
                ]
                if part
            )),
            ((18, 5), _first_non_empty(context.get("WorkPlace"), context.get("CompanyName"), "по месту требования")),
            ((23, 4), _first_non_empty(_build_exam_export(exams_by_role.get("therapist")).get("diagnosis"), "Дз: практически здоров")),
            ((23, 15), _build_exam_export(exams_by_role.get("therapist")).get("doctor", "")),
            ((24, 4), _build_exam_export(exams_by_role.get("surgeon")).get("diagnosis", "")),
            ((24, 15), _build_exam_export(exams_by_role.get("surgeon")).get("doctor", "")),
            ((25, 4), _build_exam_export(exams_by_role.get("neurologist")).get("diagnosis", "")),
            ((25, 15), _build_exam_export(exams_by_role.get("neurologist")).get("doctor", "")),
            ((27, 5), _build_exam_export(exams_by_role.get("otolaryngologist")).get("diagnosis", "")),
            ((27, 15), _build_exam_export(exams_by_role.get("otolaryngologist")).get("doctor", "")),
            ((44, 1), context.get("Conclusion", "")),
            ((48, 1), _xls_excel_date(issue_date)),
            ((50, 12), _first_non_empty(_build_exam_export(exams_by_role.get("therapist")).get("doctor"), context.get("Doctor", ""))),
            ((53, 12), _first_non_empty(_build_exam_export(exams_by_role.get("chairman")).get("doctor"), _build_exam_export(exams_by_role.get("therapist")).get("doctor"), context.get("Doctor", ""))),
        ],
    )


def _fill_eeg_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams: list[DoctorExam],
) -> None:
    eeg_exam = next((exam for exam in exams if "ээг" in f"{exam.result_text or ''} {exam.diagnosis or ''}".lower()), None)
    eeg_data = _build_exam_export(eeg_exam)
    doctor_name = _first_non_empty(eeg_data.get("doctor"), context.get("Doctor"))
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((9, 8), _xls_excel_date(encounter.encounter_date if encounter else date.today())),
            ((11, 8), context.get("ClientCalc", "")),
            ((13, 8), _age_at_date(client.birth_date, encounter.encounter_date if encounter else None)),
            ((15, 8), context.get("CardNumber", "") or context.get("ReferenceNumber", "")),
            ((17, 8), doctor_name),
        ],
    )


def _fill_chod_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    encounter: Encounter | None,
) -> None:
    issue_date = encounter.encounter_date if encounter else date.today()
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((17, 15), context.get("BlankNumber", "") or context.get("ReferenceNumber", "")),
            ((20, 3), context.get("ClientCalc", "")),
            ((21, 11), context.get("BirthDateCalc_DAY", "")),
            ((21, 16), context.get("BirthDateCalc_MONTH", "")),
            ((21, 20), context.get("BirthDateCalc_YEAR", "")),
            ((23, 3), context.get("SubjectCalc", "")),
            ((24, 5), context.get("DistrictCalc", "")),
            ((25, 4), context.get("CityCalc", "")),
            ((27, 5), context.get("StreetCalc", "")),
            ((28, 5), " ".join(part for part in [context.get("HouseNumberCalc", ""), context.get("ApartmentNumberCalc", "")] if part)),
            ((30, 16), _xls_excel_date(issue_date)),
            ((36, 12), context.get("Doctor", "")),
        ],
    )


def _restriction_text(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text or text in {"0", "нет", "false", "no", "не установлено"}:
        return "не установлено"
    return "установлено"


def _xls_blank_or_dash(value: object) -> str:
    text = str(value or "").strip()
    return text or "-"


def _driver_exam_line(exam: DoctorExam | None, fallback: str = "Противопоказания отсутствуют") -> str:
    if exam is None:
        return fallback
    data = _build_exam_export(exam)
    doctor = str(data.get("doctor") or "").strip()
    if not doctor:
        return fallback
    conclusion = _first_non_empty(
        data.get("diagnosis"),
        data.get("objective"),
        data.get("title"),
        "Противопоказания отсутствуют",
    )
    return " ".join(part for part in [doctor, conclusion] if part).strip()


def _driver_auxiliary_line(context: dict[str, str], key: str, fallback: str = "Не установлено") -> str:
    value = str(context.get(key) or "").strip()
    return value or fallback


DRIVER_XLS_CATEGORY_KEYS = ("A", "B", "C", "D", "BE", "CE", "DE", "Tm", "Tb", "M", "A1", "B1", "C1", "D1", "C1E", "D1E")
DRIVER_XLS_CATEGORY_CELLS_LEFT = (
    (10, 2),
    (10, 4),
    (10, 6),
    (10, 8),
    (10, 10),
    (10, 12),
    (10, 14),
    (10, 16),
    (10, 18),
    (10, 20),
    (10, 22),
    (10, 24),
    (10, 26),
    (10, 28),
    (10, 30),
    (10, 32),
)
DRIVER_XLS_CATEGORY_CELLS_RIGHT = (
    (10, 35),
    (10, 37),
    (10, 39),
    (10, 41),
    (10, 43),
    (10, 45),
    (10, 47),
    (10, 49),
    (10, 51),
    (10, 53),
    (10, 55),
    (10, 57),
    (10, 59),
    (10, 61),
    (10, 63),
    (10, 65),
)
DRIVER_CATEGORY_FIELD_KEYS = {
    "A": "categoryA",
    "B": "categoryB",
    "C": "categoryC",
    "D": "categoryD",
    "BE": "categoryBE",
    "CE": "categoryCE",
    "DE": "categoryDE",
    "Tm": "categoryTram",
    "Tb": "categoryTrolleybus",
    "M": "categoryM",
    "A1": "categoryA1",
    "B1": "categoryB1",
    "C1": "categoryC1",
    "D1": "categoryD1",
    "C1E": "categoryC1E",
    "D1E": "categoryD1E",
}
DRIVER_XLS_FRONT_SHEET_NAMES = ("Водительская Лицевая", "Вод.Лиц22")
DRIVER_XLS_BACK_SHEET_NAMES = ("Водительская Оборотная", "Вод.Об22", "Вод.Оборот")
DRIVER_CATEGORY_LEGACY_ALIASES = {
    "A1": "1A",
    "B1": "1B",
    "C1": "1C",
    "D1": "1D",
    "C1E": "1CE",
    "D1E": "1DE",
}
DRIVER_XML_CATEGORY_TOKENS = {
    "A": ("ACalc", "CategoryA"),
    "B": ("BCalc", "CategoryB"),
    "C": ("CCalc", "CategoryC"),
    "D": ("DCalc", "CategoryD"),
    "BE": ("BECalc", "CategoryBE"),
    "CE": ("CECalc", "CategoryCE"),
    "DE": ("DECalc", "CategoryDE"),
    "Tm": ("TmCalc", "CategoryTm"),
    "Tb": ("TbCalc", "CategoryTb"),
    "M": ("MCalc", "CategoryM"),
    "A1": ("A1Calc", "CategoryA1", "Category1A"),
    "B1": ("B1Calc", "CategoryB1", "Category1B"),
    "C1": ("C1Calc", "CategoryC1", "Category1C"),
    "D1": ("D1Calc", "CategoryD1", "Category1D"),
    "C1E": ("C1ECalc", "CategoryC1E", "Category1CE"),
    "D1E": ("D1ECalc", "CategoryD1E", "Category1DE"),
}
DRIVER_INDICATION_TOKEN_FIELDS = {
    "ManualControlCalc": ("indicationManual", ("с ручным упр", "ручн.управ", "ручн управ")),
    "AutomaticTransmissionCalc": ("indicationAutomatic", ("автоматич. трансмисс", "автоматической трансмисс", "с автоматом")),
    "ParkingSystemCalc": ("indicationAcoustic", ("акустич. парковочная", "акустической парковочной", "парковочная система")),
    "VisionTCCalc": ("indicationGlasses", ("коррекции зрения", "очки", "линзы")),
    "HearingTCCalc": ("indicationHearingAid", ("компенсации потери слуха", "слуховой аппарат")),
}
DRIVER_RESTRICTION_TOKEN_FIELDS = {
    "TCA": ("restrictionAM", ("am",)),
    "TCB": ("restrictionBBE", ("b be", "bbe")),
    "TCC": ("restrictionCCE", ("c ce", "cce")),
}
DRIVER_RESTRICTION_CATEGORY_GROUPS = {
    "TCA": {"A", "M", "A1", "B1"},
    "TCB": {"B", "BE", "B1"},
    "TCC": {"C", "CE", "D", "DE", "Tm", "Tb", "C1", "D1", "C1E", "D1E"},
}
DRIVER_XLS_BACK_STATUS_ROWS = (
    (14, "TCA"),
    (17, "TCB"),
    (20, "TCC"),
    (25, "ManualControlCalc"),
    (27, "AutomaticTransmissionCalc"),
    (29, "ParkingSystemCalc"),
    (31, "VisionTCCalc"),
    (33, "HearingTCCalc"),
)
DRIVER_XLS_BACK_STATUS_COLS = (29, 62)


def _truthy_driver_value(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text not in {"", "0", "false", "no", "нет", "не установлено", "z"}


def _driver_category_tokens(value: object) -> set[str]:
    text = str(value or "")
    raw_tokens = re.findall(r"[A-Za-zА-Яа-я0-9]+", text)
    aliases = {
        "1A": "A1",
        "1B": "B1",
        "1C": "C1",
        "1D": "D1",
        "1CE": "C1E",
        "1DE": "D1E",
        "Е": "E",
        "ВЕ": "BE",
        "СЕ": "CE",
        "ДЕ": "DE",
    }
    tokens: set[str] = set()
    for token in raw_tokens:
        normalized = aliases.get(token.upper(), token)
        normalized = {"TM": "Tm", "TB": "Tb"}.get(normalized.upper(), normalized.upper())
        if normalized in DRIVER_XLS_CATEGORY_KEYS or normalized == "E":
            tokens.add(normalized)
    if "E" in tokens:
        tokens.update({"BE", "CE", "DE"})
    return tokens


def _driver_completed_chairman(exams: list[DoctorExam]) -> DoctorExam | None:
    return next(
        (
            exam
            for exam in exams
            if str(exam.doctor_role_id or "").strip().lower() == "chairman" and bool(exam.is_completed)
        ),
        None,
    )


def _driver_latest_chairman(exams: list[DoctorExam]) -> DoctorExam | None:
    return next(
        (
            exam
            for exam in exams
            if str(exam.doctor_role_id or "").strip().lower() == "chairman"
        ),
        None,
    )


def _driver_categories_from_chairman(fields: dict) -> set[str] | None:
    exact_fields = {field_key for field_key in DRIVER_CATEGORY_FIELD_KEYS.values() if field_key in fields}
    if "categoryE" in fields:
        exact_fields.add("categoryE")
    if exact_fields:
        selected = {
            category
            for category, field_key in DRIVER_CATEGORY_FIELD_KEYS.items()
            if _truthy_driver_value(fields.get(field_key))
        }
        if _truthy_driver_value(fields.get("categoryE")) and not selected.intersection({"BE", "CE", "DE"}):
            selected.update({"BE", "CE", "DE"})
        return selected
    driver_categories = fields.get("driverCategories")
    if str(driver_categories or "").strip():
        return _driver_category_tokens(driver_categories)
    return None


def _driver_categories_from_context(context: dict[str, str], client: Client) -> set[str]:
    selected = {
        category
        for category in DRIVER_XLS_CATEGORY_KEYS
        if _truthy_driver_value(context.get(f"Category{category}") or context.get(f"{category}Calc"))
    }
    selected.update(_driver_category_tokens(client.admission_category))
    return selected


def _driver_categories_for_documents(client: Client, exams: list[DoctorExam]) -> set[str]:
    chairman = _driver_completed_chairman(exams)
    chairman_categories = _driver_categories_from_chairman(chairman.fields_json or {}) if chairman else None
    if chairman_categories is not None:
        return chairman_categories
    return _driver_category_tokens(client.admission_category)


def _driver_category_context_values(selected: set[str], true_value: str = "X", false_value: str = "") -> dict[str, str]:
    context: dict[str, str] = {}
    for category in DRIVER_XLS_CATEGORY_KEYS:
        value = true_value if category in selected else false_value
        for key in (
            f"Category{category}",
            f"Category{category}1",
            f"{category}Calc",
            f"Category{category}Calc",
        ):
            context[key] = value
        legacy_key = DRIVER_CATEGORY_LEGACY_ALIASES.get(category)
        if legacy_key:
            for key in (
                f"Category{legacy_key}",
                f"Category{legacy_key}1",
                f"{legacy_key}Calc",
                f"Category{legacy_key}Calc",
            ):
                context[key] = value
    return context


def _driver_text_contains(value: object, needles: tuple[str, ...]) -> bool:
    text = str(value or "").strip().lower().replace("ё", "е")
    return any(needle in text for needle in needles)


def _driver_field_or_text_flag(fields: dict, field_key: str, fallback_text: object, labels: tuple[str, ...]) -> bool:
    if field_key in fields:
        return _truthy_driver_value(fields.get(field_key))
    return _driver_text_contains(fallback_text, labels)


def _driver_flag_context_values(client: Client, exams: list[DoctorExam], selected: set[str] | None = None) -> dict[str, str]:
    chairman = _driver_completed_chairman(exams) or _driver_latest_chairman(exams)
    fields = (chairman.fields_json or {}) if chairman else {}
    fallback_text = client.indications or ""
    selected_categories = selected if selected is not None else _driver_categories_for_documents(client, exams)
    context: dict[str, str] = {}
    for token, (field_key, labels) in DRIVER_INDICATION_TOKEN_FIELDS.items():
        context[token] = "true" if _driver_field_or_text_flag(fields, field_key, fallback_text, labels) else "false"
    for token, (field_key, labels) in DRIVER_RESTRICTION_TOKEN_FIELDS.items():
        category_group = DRIVER_RESTRICTION_CATEGORY_GROUPS.get(token, set())
        has_category = bool(selected_categories.intersection(category_group))
        has_restriction = _driver_field_or_text_flag(fields, field_key, fallback_text, labels)
        context[token] = "X" if has_restriction or has_category else ""
    context["DriveShipCalc"] = "true" if _truthy_driver_value(fields.get("categoryBoat")) else "false"
    return context


def _driver_document_context_overrides(client: Client, exams: list[DoctorExam]) -> dict[str, str]:
    selected = _driver_categories_for_documents(client, exams)
    return {
        **_driver_category_context_values(selected),
        **_driver_flag_context_values(client, exams, selected),
    }


def _bool_text(value: bool) -> str:
    return "true" if value else "false"


def _driver_xml_context_overrides(context: dict[str, str], client: Client, exams: list[DoctorExam]) -> dict[str, str]:
    selected = _driver_categories_for_documents(client, exams)
    overrides: dict[str, str] = {}
    for category, token_names in DRIVER_XML_CATEGORY_TOKENS.items():
        value = _bool_text(category in selected)
        for token_name in token_names:
            overrides[token_name] = value

    flag_context = _driver_flag_context_values(client, exams, selected)
    overrides.update(
        {
            token_name: flag_context[token_name]
            for token_name in DRIVER_INDICATION_TOKEN_FIELDS
        }
    )
    overrides["DriveShipCalc"] = flag_context["DriveShipCalc"]
    overrides["CategoryACalc"] = _bool_text(bool(flag_context.get("TCA")))
    overrides["CategoryBCalc"] = _bool_text(bool(flag_context.get("TCB")))
    overrides["CategoryCCalc"] = _bool_text(bool(flag_context.get("TCC")))
    return overrides


def _driver_category_marks(context: dict[str, str], client: Client, exams_by_role: dict[str, DoctorExam]) -> list[str]:
    chairman = exams_by_role.get("chairman")
    chairman_categories = _driver_categories_from_chairman(chairman.fields_json or {}) if chairman and chairman.is_completed else None
    selected = chairman_categories if chairman_categories is not None else _driver_categories_from_context(context, client)
    return ["✓" if category in selected else "Z" for category in DRIVER_XLS_CATEGORY_KEYS]


def _driver_marker_style(source_book, source_sheet, row_index: int, col_index: int):
    if getattr(source_book, "is_xlsx", False):
        return None

    xf = source_book.xf_list[source_sheet.cell_xf_index(row_index, col_index)]
    source_border = xf.border

    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = "Arial"
    font.height = 320
    font.bold = False
    font.italic = False
    font.underline = xlwt.Font.UNDERLINE_NONE
    style.font = font

    alignment = xlwt.Alignment()
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    alignment.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = alignment

    return style


def _write_driver_marker_cells(source_book, source_sheet, target_sheet, cells: tuple[tuple[int, int], ...], marks: list[str]) -> None:
    for (row_index, col_index), mark in zip(cells, marks):
        target_sheet.write(row_index, col_index, mark, _driver_marker_style(source_book, source_sheet, row_index, col_index))


def _fill_driver_xls_sheets(
    source_book,
    target_book,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    driver_lines = [
        _driver_exam_line(exams_by_role.get("therapist")),
        _driver_exam_line(exams_by_role.get("ophthalmologist")),
        _driver_exam_line(exams_by_role.get("neurologist"), "не установлено"),
        _driver_exam_line(exams_by_role.get("otolaryngologist"), "не установлено"),
        _driver_auxiliary_line(context, "InstrumentalExamination"),
        _driver_auxiliary_line(context, "LaboratoryStudy"),
        _build_exam_export(exams_by_role.get("chairman")).get("doctor"),
    ]
    issue_date = encounter.encounter_date if encounter else date.today()
    front_source, front_target, _ = _sheet_pair_any(source_book, target_book, DRIVER_XLS_FRONT_SHEET_NAMES)
    if front_source and front_target:
        _write_xls_pairs(
            front_target,
            front_source,
            [
                ((15, 2), context.get("ClientCalc", "")),
                ((15, 28), context.get("ClientCalc", "")),
                ((16, 8), context.get("BirthDateCalc_DAY", "")),
                ((16, 15), context.get("BirthDateCalc_DATEMONTH", "")),
                ((16, 22), context.get("BirthDateCalc_YEAR", "")),
                ((16, 35), context.get("BirthDateCalc_DAY", "")),
                ((16, 41), context.get("BirthDateCalc_DATEMONTH", "")),
                ((16, 48), context.get("BirthDateCalc_YEAR", "")),
                ((18, 9), context.get("SubjectCalc", "")),
                ((18, 36), context.get("SubjectCalc", "")),
                ((19, 4), _xls_blank_or_dash(context.get("DistrictCalc"))),
                ((19, 31), _xls_blank_or_dash(context.get("DistrictCalc"))),
                ((20, 4), context.get("CityCalc", "")),
                ((20, 30), context.get("CityCalc", "")),
                ((21, 2), context.get("StreetCalc", "")),
                ((21, 18), context.get("HouseNumberCalc", "")),
                ((21, 30), context.get("StreetCalc", "")),
                ((21, 45), context.get("HouseNumberCalc", "")),
                ((22, 3), _xls_blank_or_dash(context.get("HouseBodyCalc"))),
                ((22, 12), context.get("ApartmentNumberCalc", "")),
                ((22, 31), _xls_blank_or_dash(context.get("HouseBodyCalc"))),
                ((22, 38), context.get("ApartmentNumberCalc", "")),
                ((23, 15), str(issue_date.day)),
                ((23, 19), context.get("VisitDate_DATEMONTH", "")),
                ((23, 23), str(issue_date.year)),
                ((23, 41), str(issue_date.day)),
                ((23, 45), context.get("VisitDate_DATEMONTH", "")),
                ((23, 49), str(issue_date.year)),
                ((28, 12), driver_lines[0]),
                ((28, 39), driver_lines[0]),
                ((30, 12), driver_lines[1]),
                ((30, 39), driver_lines[1]),
                ((35, 12), driver_lines[2]),
                ((35, 39), driver_lines[2]),
                ((37, 12), driver_lines[3]),
                ((37, 39), driver_lines[3]),
                ((39, 12), driver_lines[4]),
                ((39, 39), driver_lines[4]),
                ((41, 12), driver_lines[5]),
                ((41, 39), driver_lines[5]),
            ],
        )
        for target_coord, style_coord in [
            ((28, 12), (28, 39)),
            ((30, 12), (30, 39)),
            ((35, 12), (35, 39)),
            ((37, 12), (37, 39)),
            ((39, 12), (39, 39)),
            ((41, 12), (41, 39)),
        ]:
            _copy_xls_target_cell_style(front_target, *target_coord, *style_coord)
    back_source, back_target, _ = _sheet_pair_any(source_book, target_book, DRIVER_XLS_BACK_SHEET_NAMES)
    if back_source and back_target:
        category_marks = _driver_category_marks(context, client, exams_by_role)
        _write_driver_marker_cells(source_book, back_source, back_target, DRIVER_XLS_CATEGORY_CELLS_LEFT, category_marks)
        _write_driver_marker_cells(source_book, back_source, back_target, DRIVER_XLS_CATEGORY_CELLS_RIGHT, category_marks)
        back_cells = [
            ((36, 8), driver_lines[6]),
            ((36, 41), driver_lines[6]),
        ]
        for row_index, context_key in DRIVER_XLS_BACK_STATUS_ROWS:
            status_value = _restriction_text(context.get(context_key))
            for col_index in DRIVER_XLS_BACK_STATUS_COLS:
                back_cells.append(((row_index, col_index), status_value))
        _write_xls_pairs(back_target, back_source, back_cells)


def _fill_tractor_xls_sheets(source_book, target_book, exams_by_role: dict[str, DoctorExam]) -> None:
    tractor_lines = [
        _exam_conclusion_line(exams_by_role.get("therapist")),
        _exam_conclusion_line(exams_by_role.get("ophthalmologist")),
        _exam_conclusion_line(exams_by_role.get("neurologist")),
        _exam_conclusion_line(exams_by_role.get("otolaryngologist")),
    ]
    front_source, front_target, _ = _sheet_pair(source_book, target_book, "Тракторная Лицевая")
    if front_source and front_target:
        _write_xls_pairs(
            front_target,
            front_source,
            [
                ((29, 12), tractor_lines[0]),
                ((29, 39), tractor_lines[0]),
                ((31, 12), tractor_lines[1]),
                ((31, 39), tractor_lines[1]),
                ((35, 12), tractor_lines[2]),
                ((35, 39), tractor_lines[2]),
                ((37, 12), tractor_lines[3]),
                ((37, 39), tractor_lines[3]),
            ],
        )
    back_source, back_target, _ = _sheet_pair(source_book, target_book, "Тракторная оборотная")
    if back_source and back_target:
        signer = _first_non_empty(_build_exam_export(exams_by_role.get("chairman")).get("doctor"), _build_exam_export(exams_by_role.get("therapist")).get("doctor"))
        _write_xls_pairs(
            back_target,
            back_source,
            [
                ((36, 5), signer),
                ((36, 25), signer),
            ],
        )


def _new_xls_context_value(context: dict[str, str], key: str) -> str:
    value = str(context.get(key, "") or "").strip()
    if value.casefold() in {"не указано", "здоров", "врач", "администратор системы"}:
        return ""
    return value


def _new_xls_exam_value(exam: DoctorExam | None, *field_names: str) -> str:
    if exam is None:
        return ""
    fields = exam.fields_json or {}
    field_value = _exam_field(fields, *field_names)
    if field_value:
        return field_value
    return _first_non_empty(exam.result_text, exam.diagnosis)


def _new_xls_exam_doctor(exam: DoctorExam | None) -> str:
    return str(getattr(exam, "doctor_name", "") or "").strip()


def _new_xls_gsu_secondary_page_text(value: str) -> str:
    # The source form centers these fields across a selection whose printable
    # area starts very close to the first glyph. Keep a small internal indent
    # so desktop Excel does not clip the first letters at the page edge.
    return f"        {value}" if value else ""


def _new_xls_date_text(value: date | datetime | str | None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _new_xls_exam_date(exam: DoctorExam | None, encounter: Encounter | None) -> object:
    if exam is not None and exam.completed_at:
        return _new_xls_date_text(exam.completed_at)
    return _new_xls_date_text(encounter.encounter_date if encounter else None)


def _new_xls_signer(
    context: dict[str, str],
    exams_by_role: dict[str, DoctorExam],
) -> str:
    return _first_non_empty(
        _new_xls_exam_doctor(exams_by_role.get("chairman")),
        _new_xls_exam_doctor(exams_by_role.get("therapist")),
        _new_xls_context_value(context, "Doctor"),
    )


def _write_xls_characters(
    target_sheet,
    source_sheet,
    coordinates: tuple[tuple[int, int], ...],
    value: object,
    *,
    digits_only: bool = False,
) -> None:
    text = str(value or "").strip()
    if digits_only:
        text = re.sub(r"\D", "", text)
    for index, (row_index, col_index) in enumerate(coordinates):
        character = text[index] if index < len(text) else ""
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, character)


def _new_xls_medical_values(
    context: dict[str, str],
    exams_by_role: dict[str, DoctorExam],
) -> dict[str, str]:
    chairman = exams_by_role.get("chairman")
    therapist = exams_by_role.get("therapist")
    source_exam = chairman or therapist
    fields = source_exam.fields_json if source_exam is not None else {}
    diagnosis = _first_non_empty(
        str(getattr(source_exam, "diagnosis", "") or "").strip(),
        _exam_field(fields or {}, "diagnosis", "diagnosisShort", "diagnosisText", "diagnoz"),
        _new_xls_context_value(context, "Diagnosis"),
    )
    mkb10 = _first_non_empty(
        _exam_field(fields or {}, "mkb10", "MKB10", "icd10", "diagnosisCode"),
        _new_xls_context_value(context, "MKB10"),
    )
    diagnosis_with_code = diagnosis
    if mkb10 and mkb10.casefold() not in diagnosis.casefold():
        diagnosis_with_code = " ".join(part for part in [diagnosis, f"МКБ-10: {mkb10}"] if part)
    return {
        "diagnosis": diagnosis,
        "mkb10": mkb10,
        "diagnosis_with_code": diagnosis_with_code,
        "complaints": _exam_field(fields or {}, "complaints", "complaint", "complaintsText"),
        "anamnesis": _exam_field(
            fields or {},
            "anamnesis",
            "anamnesisText",
            "history",
            "anamnesisVitae",
        ),
        "results": _first_non_empty(
            _exam_field(
                fields or {},
                "objective",
                "objectiveData",
                "objectiveText",
                "research",
                "researchResults",
            ),
            str(getattr(source_exam, "result_text", "") or "").strip(),
        ),
        "conclusion": _first_non_empty(
            _exam_field(fields or {}, "conclusionText", "issuedConclusion", "conclusion", "result"),
            str(getattr(source_exam, "result_text", "") or "").strip(),
            str(getattr(source_exam, "diagnosis", "") or "").strip(),
            _new_xls_context_value(context, "Conclusion"),
        ),
    }


def _fill_new_gsu_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    psychiatrist = exams_by_role.get("psychiatrist")
    narcologist = exams_by_role.get("psychiatrist-narcologist")
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            (
                (4, 38),
                _new_xls_gsu_secondary_page_text(
                    _new_xls_exam_value(narcologist, "conclusion", "result")
                ),
            ),
            ((6, 39), "Врач: психиатр-нарколог" if _new_xls_exam_doctor(narcologist) else ""),
            ((6, 52), _new_xls_exam_doctor(narcologist)),
            ((7, 47), _new_xls_exam_value(narcologist, "objective", "researchResults")),
            ((18, 9), _new_xls_date_text(encounter.encounter_date if encounter else None)),
            (
                (18, 19),
                _first_non_empty(
                    context.get("BlankFullNumber"),
                    context.get("BlankNumber"),
                    context.get("ReferenceNumber"),
                ),
            ),
            (
                (19, 38),
                _new_xls_gsu_secondary_page_text(
                    _new_xls_exam_value(psychiatrist, "conclusion", "result")
                ),
            ),
            ((21, 39), "Врач: психиатр" if _new_xls_exam_doctor(psychiatrist) else ""),
            ((21, 52), _new_xls_exam_doctor(psychiatrist)),
            ((22, 47), _new_xls_exam_value(psychiatrist, "objective", "researchResults")),
            ((27, 11), context.get("ClientCalc", "")),
            ((30, 14), context.get("SexFull") or context.get("SexCalc", "")),
            ((31, 10), _new_xls_date_text(client.birth_date)),
            ((33, 2), context.get("AddressCalc", "")),
            ((38, 26), _new_xls_exam_doctor(exams_by_role.get("therapist"))),
            ((41, 26), _new_xls_signer(context, exams_by_role)),
        ],
    )


def _fill_new_070_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    medical = _new_xls_medical_values(context, exams_by_role)
    source_exam = exams_by_role.get("chairman") or exams_by_role.get("therapist")
    fields = source_exam.fields_json if source_exam is not None else {}
    sanatorium = _first_non_empty(
        _exam_field(fields or {}, "sanatorium", "sanatoriumName", "treatmentOrganization"),
        _new_xls_context_value(context, "Sanatorium"),
    )
    preferred = _first_non_empty(
        _exam_field(fields or {}, "preferredTreatmentPlace", "preferredPlace"),
        sanatorium,
    )
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            (
                (12, 20),
                _first_non_empty(
                    context.get("BlankFullNumber"),
                    context.get("BlankNumber"),
                    context.get("ReferenceNumber"),
                ),
            ),
            ((14, 12), _new_xls_date_text(encounter.encounter_date if encounter else None)),
            ((18, 15), context.get("ClientCalc", "")),
            ((19, 7), _new_xls_date_text(client.birth_date)),
            ((19, 27), context.get("SexFull") or context.get("SexCalc", "")),
            ((21, 1), context.get("AddressCalc", "")),
            ((29, 20), _new_xls_context_value(context, "DisabilityCategoryCode")),
            ((29, 21), _new_xls_context_value(context, "BenefitCategoryCode")),
            ((33, 20), _new_xls_context_value(context, "InsuranceKind")),
            ((36, 3), context.get("DocumentSeries", "")),
            ((36, 8), context.get("DocumentNumber", "")),
            ((36, 20), context.get("DocumentDate", "")),
            ((36, 33), context.get("Phone", "")),
            ((39, 11), sanatorium),
            ((42, 1), medical["diagnosis_with_code"]),
            ((50, 10), _exam_field(fields or {}, "treatmentType", "treatmentMode")),
            ((52, 7), preferred),
            ((56, 22), _new_xls_exam_doctor(exams_by_role.get("therapist"))),
            ((57, 22), _new_xls_signer(context, exams_by_role)),
        ],
    )
    _write_xls_characters(
        target_sheet,
        source_sheet,
        tuple((25, col_index) for col_index in range(15, 31)),
        context.get("PolisOMS", ""),
        digits_only=True,
    )
    _write_xls_characters(
        target_sheet,
        source_sheet,
        tuple((37, col_index) for col_index in range(16, 30)),
        context.get("SNILS", ""),
    )


def _fill_new_072_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    medical = _new_xls_medical_values(context, exams_by_role)
    source_exam = exams_by_role.get("chairman") or exams_by_role.get("therapist")
    fields = source_exam.fields_json if source_exam is not None else {}
    sanatorium = _first_non_empty(
        _exam_field(fields or {}, "sanatorium", "sanatoriumName", "treatmentOrganization"),
        _new_xls_context_value(context, "Sanatorium"),
    )
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            (
                (12, 24),
                _first_non_empty(
                    context.get("BlankFullNumber"),
                    context.get("BlankNumber"),
                    context.get("ReferenceNumber"),
                ),
            ),
            ((13, 12), _new_xls_date_text(encounter.encounter_date if encounter else None)),
            ((15, 14), context.get("ClientCalc", "")),
            ((16, 7), _new_xls_date_text(client.birth_date)),
            ((16, 24), context.get("SexFull") or context.get("SexCalc", "")),
            ((18, 1), context.get("AddressCalc", "")),
            ((25, 20), _new_xls_context_value(context, "BenefitCategoryCode")),
            ((29, 20), _new_xls_context_value(context, "InsuranceKind")),
            ((32, 3), context.get("DocumentSeries", "")),
            ((32, 8), context.get("DocumentNumber", "")),
            ((32, 20), context.get("DocumentDate", "")),
            ((32, 33), context.get("Phone", "")),
            ((39, 14), sanatorium),
            ((42, 14), context.get("ClientCalc", "")),
            ((45, 1), medical["diagnosis_with_code"]),
            ((56, 1), medical["complaints"]),
            ((58, 1), medical["anamnesis"]),
            ((62, 1), medical["results"]),
            ((68, 1), medical["diagnosis_with_code"]),
            ((75, 1), _exam_field(fields or {}, "additionalInformation", "additionalInfo")),
            ((78, 1), _exam_field(fields or {}, "disabilityCause", "disabilityDiagnosis")),
            ((83, 17), sanatorium),
            ((84, 28), _exam_field(fields or {}, "treatmentType", "treatmentMode")),
            ((87, 16), _exam_field(fields or {}, "voucherNumber", "permitNumber")),
            ((89, 1), _new_xls_exam_doctor(source_exam)),
            ((90, 28), _new_xls_signer(context, exams_by_role)),
        ],
    )
    _write_xls_characters(
        target_sheet,
        source_sheet,
        tuple((21, col_index) for col_index in range(22, 38)),
        context.get("PolisOMS", ""),
        digits_only=True,
    )
    _write_xls_characters(
        target_sheet,
        source_sheet,
        tuple((33, col_index) for col_index in range(14, 28)),
        context.get("SNILS", ""),
    )


def _fill_new_sport_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    chairman = exams_by_role.get("chairman")
    therapist = exams_by_role.get("therapist")
    source_exam = chairman or therapist
    fields = source_exam.fields_json if source_exam is not None else {}
    issue_date = encounter.encounter_date if encounter else None
    sport_type = _first_non_empty(
        _exam_field(fields or {}, "sportType", "sport", "sports", "discipline"),
        _new_xls_context_value(context, "SportType"),
    )
    ekg = _first_non_empty(
        _exam_field(fields or {}, "ekgConclusion", "EKGConclusion", "ecgConclusion", "ekg", "EKG", "ecg"),
        _new_xls_context_value(context, "SportEkgConclusion"),
        _new_xls_context_value(context, "SportEkg"),
    )
    conclusion = _first_non_empty(
        _exam_field(fields or {}, "conclusionText", "sportConclusionText", "issuedConclusion", "conclusion", "result"),
        str(getattr(source_exam, "result_text", "") or "").strip(),
        _new_xls_context_value(context, "SportConclusion"),
    )
    valid_until = _first_non_empty(
        _exam_field(fields or {}, "validUntil", "validThrough", "sportValidUntil"),
        context.get("PoolValidUntil"),
    )
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            (
                (1, 13),
                _first_non_empty(
                    context.get("BlankFullNumber"),
                    context.get("BlankNumber"),
                    context.get("ReferenceNumber"),
                ),
            ),
            ((12, 8), _new_xls_date_text(issue_date)),
            ((13, 3), context.get("ClientCalc", "")),
            ((13, 13), _new_xls_date_text(client.birth_date)),
            ((15, 5), _new_xls_date_text(issue_date)),
            ((19, 7), sport_type),
            ((22, 2), ekg),
            ((23, 3), conclusion),
            ((27, 11), _new_xls_signer(context, exams_by_role)),
            ((29, 7), valid_until),
        ],
    )


def _fill_new_gostaina_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    narcologist = exams_by_role.get("psychiatrist-narcologist")
    psychiatrist = exams_by_role.get("psychiatrist")
    neurologist = exams_by_role.get("neurologist")
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((16, 6), _new_xls_date_text(encounter.encounter_date if encounter else None)),
            (
                (16, 11),
                _first_non_empty(
                    context.get("BlankFullNumber"),
                    context.get("BlankNumber"),
                    context.get("ReferenceNumber"),
                ),
            ),
            ((22, 2), context.get("ClientCalc", "")),
            ((24, 4), _new_xls_date_text(client.birth_date)),
            ((26, 4), context.get("SexFull") or context.get("SexCalc", "")),
            ((29, 0), context.get("AddressCalc", "")),
            ((36, 9), _new_xls_exam_date(narcologist, encounter)),
            ((36, 15), _new_xls_exam_doctor(narcologist)),
            ((38, 9), _new_xls_exam_date(psychiatrist, encounter)),
            ((38, 15), _new_xls_exam_doctor(psychiatrist)),
            ((40, 9), _new_xls_exam_date(neurologist, encounter)),
            ((40, 15), _new_xls_exam_doctor(neurologist)),
            ((47, 11), _new_xls_signer(context, exams_by_role)),
            ((51, 15), _new_xls_exam_doctor(neurologist)),
            ((53, 15), _new_xls_exam_doctor(narcologist)),
            ((55, 15), _new_xls_exam_doctor(psychiatrist)),
        ],
    )


def _fill_new_tractor_back_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    exams_by_role: dict[str, DoctorExam],
) -> None:
    roles = (
        "therapist",
        "ophthalmologist",
        "neurologist",
        "otolaryngologist",
        "surgeon",
        "psychiatrist",
        "psychiatrist-narcologist",
        "gynecologist",
        "dermatologist",
    )
    rows = (9, 11, 14, 17, 19, 20, 21, 22, 23)
    pairs: list[tuple[tuple[int, int], object]] = []
    for row_index, role_id in zip(rows, roles):
        value = _new_xls_exam_value(
            exams_by_role.get(role_id),
            "conclusion",
            "result",
            "diagnosis",
        )
        pairs.extend([((row_index, 18), value), ((row_index, 38), value)])
    signer = _new_xls_signer(context, exams_by_role)
    pairs.extend([((36, 5), signer), ((36, 25), signer)])
    _write_xls_pairs(target_sheet, source_sheet, pairs)


def _fill_new_tractor_front_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    issue_date = encounter.encounter_date if encounter else None
    address = {
        "subject": context.get("SubjectCalc", ""),
        "district": context.get("DistrictCalc", ""),
        "city": context.get("CityCalc", ""),
        "street": context.get("StreetCalc", ""),
        "house": context.get("HouseNumberCalc", ""),
        "body": context.get("HouseBodyCalc", ""),
        "apartment": context.get("ApartmentNumberCalc", ""),
    }
    if not any(address.values()) and context.get("AddressCalc"):
        parsed_address = _split_address(context.get("AddressCalc", ""))
        address = {
            "subject": parsed_address.get("subject", ""),
            "district": parsed_address.get("district", ""),
            "city": parsed_address.get("city", ""),
            "street": parsed_address.get("street", ""),
            "house": parsed_address.get("house", ""),
            "body": parsed_address.get("body", ""),
            "apartment": parsed_address.get("apartment", ""),
        }
    has_address = any(address.values())
    blank_number = _first_non_empty(
        context.get("BlankNumber"),
        context.get("BlankFullNumber"),
        context.get("ReferenceNumber"),
    )
    exam_lines = []
    for role_id in ("therapist", "ophthalmologist", "neurologist", "otolaryngologist"):
        exam = exams_by_role.get(role_id)
        exam_lines.append(_exam_conclusion_line(exam) if exam is not None else "")

    values = [
        ((7, 3), blank_number),
        ((7, 30), blank_number),
        ((14, 2), context.get("ClientCalc", "")),
        ((14, 28), context.get("ClientCalc", "")),
        ((15, 8), context.get("BirthDateCalc_DAY", "")),
        ((15, 15), context.get("BirthDateCalc_DATEMONTH", "")),
        ((15, 22), context.get("BirthDateCalc_YEAR", "")),
        ((15, 35), context.get("BirthDateCalc_DAY", "")),
        ((15, 41), context.get("BirthDateCalc_DATEMONTH", "")),
        ((15, 48), context.get("BirthDateCalc_YEAR", "")),
        ((17, 12), address["subject"]),
        ((17, 38), address["subject"]),
        ((18, 4), _xls_blank_or_dash(address["district"]) if has_address else ""),
        ((18, 31), _xls_blank_or_dash(address["district"]) if has_address else ""),
        ((19, 6), address["city"]),
        ((19, 32), address["city"]),
        ((20, 2), address["street"]),
        ((20, 30), address["street"]),
        ((21, 2), address["house"]),
        ((21, 29), address["house"]),
        ((22, 2), _xls_blank_or_dash(address["body"]) if has_address else ""),
        ((22, 9), address["apartment"]),
        ((22, 29), _xls_blank_or_dash(address["body"]) if has_address else ""),
        ((22, 36), address["apartment"]),
        ((23, 12), context.get("SNILS", "")),
        ((23, 39), context.get("SNILS", "")),
        ((26, 2), str(issue_date.day) if issue_date else ""),
        ((26, 11), MONTH_NAMES.get(issue_date.month, "") if issue_date else ""),
        ((26, 21), str(issue_date.year) if issue_date else ""),
        ((26, 29), str(issue_date.day) if issue_date else ""),
        ((26, 38), MONTH_NAMES.get(issue_date.month, "") if issue_date else ""),
        ((26, 47), str(issue_date.year) if issue_date else ""),
        ((29, 12), exam_lines[0]),
        ((29, 39), exam_lines[0]),
        ((31, 12), exam_lines[1]),
        ((31, 39), exam_lines[1]),
        ((35, 12), exam_lines[2]),
        ((35, 39), exam_lines[2]),
        ((37, 12), exam_lines[3]),
        ((37, 39), exam_lines[3]),
    ]
    _write_xls_pairs(target_sheet, source_sheet, values)


def _fill_new_gims_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    issue_date = encounter.encounter_date if encounter else None
    address = {
        "city": context.get("CityCalc", ""),
        "district": context.get("DistrictCalc", ""),
        "street": context.get("StreetCalc", ""),
        "house": context.get("HouseNumberCalc", ""),
        "body": context.get("HouseBodyCalc", ""),
        "apartment": context.get("ApartmentNumberCalc", ""),
    }
    if not any(address.values()) and context.get("AddressCalc"):
        address = _split_address(context.get("AddressCalc", ""))
    blank_number = _first_non_empty(
        context.get("BlankNumber"),
        context.get("BlankFullNumber"),
    )
    signer = _new_xls_signer(context, exams_by_role)
    values = [
        ((7, 3), blank_number),
        ((7, 30), blank_number),
        ((14, 2), context.get("ClientCalc", "")),
        ((14, 28), context.get("ClientCalc", "")),
        ((15, 14), context.get("BirthDateCalc_DAY", "")),
        ((15, 17), context.get("BirthDateCalc_DATEMONTH", "")),
        ((15, 21), context.get("BirthDateCalc_YEAR", "")),
        ((15, 39), context.get("BirthDateCalc_DAY", "")),
        ((15, 42), context.get("BirthDateCalc_DATEMONTH", "")),
        ((15, 46), context.get("BirthDateCalc_YEAR", "")),
        ((17, 3), context.get("SNILS", "")),
        ((17, 29), context.get("SNILS", "")),
        ((18, 17), address.get("city", "")),
        ((18, 43), address.get("city", "")),
        ((19, 4), address.get("district", "")),
        ((19, 32), address.get("district", "")),
        ((20, 7), address.get("city", "")),
        ((20, 35), address.get("city", "")),
        ((21, 4), address.get("street", "")),
        ((21, 20), address.get("house", "")),
        ((21, 31), address.get("street", "")),
        ((21, 47), address.get("house", "")),
        ((22, 8), address.get("body", "")),
        ((22, 17), address.get("apartment", "")),
        ((22, 34), address.get("body", "")),
        ((22, 41), address.get("apartment", "")),
        ((34, 15), issue_date.day if issue_date else ""),
        ((34, 19), issue_date.strftime("%m") if issue_date else ""),
        ((34, 23), issue_date.year if issue_date else ""),
        ((34, 39), issue_date.day if issue_date else ""),
        ((34, 44), issue_date.strftime("%m") if issue_date else ""),
        ((34, 48), issue_date.year if issue_date else ""),
        ((36, 9), signer),
        ((36, 31), signer),
    ]
    _write_xls_pairs(target_sheet, source_sheet, values)


def _fill_new_xls_sheets(
    source_book,
    target_book,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    for sheet_name, spec in NEW_XLS_TEMPLATE_BY_SHEET.items():
        source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, sheet_name)
        if source_sheet is None or target_sheet is None:
            continue
        _clear_xls_cells(target_sheet, source_sheet, list(spec.dynamic_cells))
        if sheet_name == "ГС":
            _fill_new_gsu_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "CKK":
            _fill_new_070_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "CKK72":
            _fill_new_072_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "Спорт":
            _fill_new_sport_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "ГТ":
            _fill_new_gostaina_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "Тр.Лиц":
            _fill_new_tractor_front_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)
        elif sheet_name == "Тр.Об":
            _fill_new_tractor_back_xls_sheet(source_sheet, target_sheet, context, exams_by_role)
        elif sheet_name == "Суда":
            _fill_new_gims_xls_sheet(source_sheet, target_sheet, context, encounter, exams_by_role)


def _fill_amb_opo_xls_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams_by_role: dict[str, DoctorExam],
) -> None:
    issue_date = encounter.encounter_date if encounter else date.today()
    work_place = ", ".join(part for part in [context.get("CompanyName", ""), context.get("Post", "")] if part and part != "не указано")
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((2, 35), context.get("ReferenceNumber", "")),
            ((16, 54), _xls_excel_date(issue_date)),
            ((17, 43), context.get("ClientCalc", "")),
            ((18, 35), context.get("SexCalc", "")),
            ((18, 48), _xls_excel_date(client.birth_date)),
            ((19, 54), context.get("CityCalc", "")),
            ((20, 35), context.get("DistrictCalc", "")),
            ((20, 49), context.get("CityCalc", "")),
            ((21, 40), context.get("StreetCalc", "")),
            ((22, 35), context.get("AddressCalc", "")),
            ((22, 56), context.get("Phone", "")),
            ((24, 55), context.get("SNILS", "")),
            ((26, 48), client.document_type or "Паспорт РФ"),
            ((26, 56), context.get("DocumentSeries", "")),
            ((26, 59), context.get("DocumentNumber", "")),
            ((38, 44), work_place),
            ((72, 11), _exam_export_with_client_doctor(exams_by_role.get("therapist"), client, "therapist").get("doctor", "")),
            ((72, 42), _exam_export_with_client_doctor(exams_by_role.get("psychiatrist"), client, "psychiatrist").get("doctor", "")),
            ((94, 11), _exam_export_with_client_doctor(exams_by_role.get("neurologist"), client, "neurologist").get("doctor", "")),
            ((94, 42), _exam_export_with_client_doctor(exams_by_role.get("otolaryngologist"), client, "otolaryngologist").get("doctor", "")),
        ],
    )


def _fill_journal_344_sheet(
    source_sheet,
    target_sheet,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
) -> None:
    issue_date = encounter.encounter_date if encounter else date.today()
    _write_xls_pairs(
        target_sheet,
        source_sheet,
        [
            ((7, 0), 1),
            ((7, 1), _xls_excel_date(issue_date)),
            ((7, 2), context.get("BlankNumber", "") or context.get("ReferenceNumber", "")),
            ((7, 3), context.get("ClientCalc", "")),
            ((7, 4), _xls_excel_date(client.birth_date)),
            ((7, 5), context.get("Conclusion", "")),
            ((7, 6), ""),
            ((7, 7), ""),
        ],
    )


def _fill_prof_extract_doctor_rows(
    source_book,
    source_sheet,
    target_sheet,
    exams_by_role: dict[str, DoctorExam],
    encounter: Encounter | None,
    client: Client | None,
) -> None:
    pairs: list[tuple[tuple[int, int], object]] = []
    for _, _, row_index in PROF_EXTRACT_DOCTOR_ROWS:
        pairs.extend(
            [
                ((row_index, PROF_EXTRACT_SEQUENCE_COL), ""),
                ((row_index, PROF_EXTRACT_DOCTOR_COL), ""),
                ((row_index, PROF_EXTRACT_DATE_COL), ""),
                ((row_index, PROF_EXTRACT_CONCLUSION_COL), ""),
            ]
        )
    row_values = _prof_extract_doctor_row_values(exams_by_role, encounter, client)
    doctor_style_by_row = {
        row_index: _xls_shrink_to_fit_style(source_book, source_sheet, row_index, PROF_EXTRACT_DOCTOR_COL)
        for row_index in {row_index for row_index, _, _, _ in row_values}
    }
    for sequence_number, (row_index, doctor_line, completed_date, conclusion) in enumerate(row_values, start=1):
        pairs.extend(
            [
                ((row_index, PROF_EXTRACT_SEQUENCE_COL), sequence_number),
                ((row_index, PROF_EXTRACT_DATE_COL), _prof_xls_display_date(completed_date)),
                ((row_index, PROF_EXTRACT_CONCLUSION_COL), conclusion),
            ]
        )
    _write_xls_pairs(target_sheet, source_sheet, pairs)
    for row_index, doctor_line, _, _ in row_values:
        _write_xls_cell(
            target_sheet,
            source_sheet,
            row_index,
            PROF_EXTRACT_DOCTOR_COL,
            doctor_line,
            doctor_style_by_row[row_index],
        )


def _find_prof_amb_sheet_index(source_book) -> int | None:
    for index, sheet_name in enumerate(source_book.sheet_names()):
        normalized = str(sheet_name or "").strip().lower().replace("!", "").strip()
        if normalized in {"амб", "àìá"}:
            return index
    return None


def _generate_prof_amb_xls(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams: list[DoctorExam],
    print_variant: str | None = None,
) -> None:
    source_book = xlrd.open_workbook(file_contents=template_path.read_bytes(), formatting_info=True)
    amb_index = _find_prof_amb_sheet_index(source_book)
    if amb_index is None:
        raise ValueError("В шаблоне не найден лист амбулаторной карты")
    source_sheet = source_book.sheet_by_index(amb_index)
    target_book = copy_xls_workbook(source_book)
    target_sheet = target_book.get_sheet(amb_index)
    target_book._Workbook__active_sheet = amb_index

    exams_by_role = _exam_map(exams)
    _fill_driver_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)

    address = context.get("AddressCalc", "")
    city = context.get("CityCalc", "")
    district = context.get("DistrictCalc", "")
    street = context.get("StreetCalc", "") or address
    oms_series, oms_number = _split_policy(context.get("PolisOMS", ""))
    passport_series, passport_number = _split_document(context.get("DocumentSeries", ""), context.get("DocumentNumber", ""))
    visit_date = encounter.encounter_date if encounter else None
    work_place = ", ".join(part for part in [context.get("CompanyName", ""), context.get("Post", "")] if part and part != "не указано")

    header_values: list[tuple[tuple[int, int], object]] = [
        ((15, 54), context.get("ReferenceNumber", "")),
        ((16, 47), _xls_excel_date(visit_date)),
        ((17, 43), context.get("ClientCalc", "")),
        ((18, 35), context.get("SexCalc", "")),
        ((18, 48), _xls_excel_date(client.birth_date)),
        ((19, 54), context.get("SubjectCalc", "")),
        ((20, 35), district),
        ((20, 50), city),
        ((21, 40), city),
        ((22, 35), street),
        ((22, 56), context.get("Phone", "")),
        ((23, 48), 2 if _is_rural_address(address, city, district) else 1),
        ((24, 35), oms_series),
        ((24, 46), oms_number),
        ((24, 55), context.get("SNILS", "")),
        ((26, 48), "Паспорт РФ" if passport_series or passport_number else ""),
        ((26, 56), passport_series),
        ((26, 59), passport_number),
        ((38, 44), work_place),
        ((47, 40), context.get("BloodGroup", "")),
        ((47, 54), context.get("RhFactor", "") or context.get("RhesusFactor", "")),
        ((48, 44), context.get("Allergies", "")),
    ]
    for (row_index, col_index), value in header_values:
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, value)

    _clear_xls_cells(
        target_sheet,
        source_sheet,
        [
            (30, 1),
            (31, 14),
            (33, 1),
            (40, 1),
            (43, 1),
            (45, 18),
            (45, 26),
            (47, 18),
            (47, 26),
            (49, 18),
            (49, 26),
        ],
    )

    empty_exam_block = {
        "date": "",
        "title": "",
        "complaints": "",
        "anamnesis": "",
        "objective": "",
        "diagnosis": "",
        "doctor": "",
    }
    for block in PROF_AMB_EXAM_BLOCKS:
        _fill_exam_block(target_sheet, source_sheet, empty_exam_block, source_book=source_book, **block)
    exam_block_values = _prof_amb_exam_block_values(exams_by_role, encounter, client)
    for block in PROF_AMB_EXAM_BLOCKS[len(exam_block_values) :]:
        _clear_prof_amb_exam_block(target_sheet, source_sheet, block)
    for block, (_, exam_data) in zip(PROF_AMB_EXAM_BLOCKS, exam_block_values):
        _fill_exam_block(target_sheet, source_sheet, exam_data, source_book=source_book, **block)

    pz2_source, pz2_target, _ = _sheet_pair(source_book, target_book, "ПЗ2")
    if pz2_source and pz2_target:
        _fill_prof_extract_doctor_rows(source_book, pz2_source, pz2_target, exams_by_role, encounter, client)

    chod_source, chod_target, _ = _sheet_pair(source_book, target_book, "ЧОД")
    if chod_source and chod_target:
        _fill_chod_xls_sheet(chod_source, chod_target, context, encounter)

    _apply_xls_auto_markers(source_book, target_book, context, client, encounter, exams_by_role)
    _apply_print_variant_to_xls_workbook(target_book, print_variant)
    target_book.save(str(output_path))


def _generate_prof_amb_xlsx(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams: list[DoctorExam],
    print_variant: str | None = None,
) -> None:
    source_book = _XlsxBookAdapter(load_workbook(template_path, data_only=False))
    amb_index = _find_prof_amb_sheet_index(source_book)
    if amb_index is None:
        raise ValueError("В шаблоне не найден лист амбулаторной карты")

    target_book = _XlsxBookAdapter(load_workbook(template_path, data_only=False))
    source_sheet = source_book.sheet_by_index(amb_index)
    target_sheet = target_book.get_sheet(amb_index)

    exams_by_role = _exam_map(exams)
    _fill_driver_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)

    address = context.get("AddressCalc", "")
    city = context.get("CityCalc", "")
    district = context.get("DistrictCalc", "")
    street = context.get("StreetCalc", "") or address
    oms_series, oms_number = _split_policy(context.get("PolisOMS", ""))
    passport_series, passport_number = _split_document(context.get("DocumentSeries", ""), context.get("DocumentNumber", ""))
    visit_date = encounter.encounter_date if encounter else None
    work_place = ", ".join(part for part in [context.get("CompanyName", ""), context.get("Post", "")] if part and part != "не указано")

    header_values: list[tuple[tuple[int, int], object]] = [
        ((15, 54), context.get("ReferenceNumber", "")),
        ((16, 47), _xls_excel_date(visit_date)),
        ((17, 43), context.get("ClientCalc", "")),
        ((18, 35), context.get("SexCalc", "")),
        ((18, 48), _xls_excel_date(client.birth_date)),
        ((19, 54), context.get("SubjectCalc", "")),
        ((20, 35), district),
        ((20, 50), city),
        ((21, 40), city),
        ((22, 35), street),
        ((22, 56), context.get("Phone", "")),
        ((23, 48), 2 if _is_rural_address(address, city, district) else 1),
        ((24, 35), oms_series),
        ((24, 46), oms_number),
        ((24, 55), context.get("SNILS", "")),
        ((26, 48), "Паспорт РФ" if passport_series or passport_number else ""),
        ((26, 56), passport_series),
        ((26, 59), passport_number),
        ((38, 44), work_place),
        ((47, 40), context.get("BloodGroup", "")),
        ((47, 54), context.get("RhFactor", "") or context.get("RhesusFactor", "")),
        ((48, 44), context.get("Allergies", "")),
    ]
    for (row_index, col_index), value in header_values:
        _write_xls_cell(target_sheet, source_sheet, row_index, col_index, value)

    _clear_xls_cells(
        target_sheet,
        source_sheet,
        [
            (30, 1),
            (31, 14),
            (33, 1),
            (40, 1),
            (43, 1),
            (45, 18),
            (45, 26),
            (47, 18),
            (47, 26),
            (49, 18),
            (49, 26),
        ],
    )

    empty_exam_block = {
        "date": "",
        "title": "",
        "complaints": "",
        "anamnesis": "",
        "objective": "",
        "diagnosis": "",
        "doctor": "",
    }
    for block in PROF_AMB_EXAM_BLOCKS:
        _fill_exam_block(target_sheet, source_sheet, empty_exam_block, source_book=source_book, **block)
    exam_block_values = _prof_amb_exam_block_values(exams_by_role, encounter, client)
    for block in PROF_AMB_EXAM_BLOCKS[len(exam_block_values) :]:
        _clear_prof_amb_exam_block(target_sheet, source_sheet, block)
    for block, (_, exam_data) in zip(PROF_AMB_EXAM_BLOCKS, exam_block_values):
        _fill_exam_block(target_sheet, source_sheet, exam_data, source_book=source_book, **block)

    pz2_source, pz2_target, _ = _sheet_pair(source_book, target_book, "ПЗ2")
    if pz2_source and pz2_target:
        _fill_prof_extract_doctor_rows(source_book, pz2_source, pz2_target, exams_by_role, encounter, client)

    chod_source, chod_target, _ = _sheet_pair(source_book, target_book, "ЧОД")
    if chod_source and chod_target:
        _fill_chod_xls_sheet(chod_source, chod_target, context, encounter)

    _apply_xls_auto_markers(source_book, target_book, context, client, encounter, exams_by_role)
    _apply_print_variant_to_xls_workbook(target_book, print_variant)
    target_book.save(output_path)


def _apply_print_variant_to_xls_workbook(target_book, print_variant: str | None) -> None:
    variant = str(print_variant or "").strip().lower()
    if not variant:
        return

    sheets_by_variant = {
        "driver_front": DRIVER_XLS_FRONT_SHEET_NAMES,
        "driver_back": DRIVER_XLS_BACK_SHEET_NAMES,
        "tractor_front": ("Тр.Лиц", "Тракторная Лицевая"),
        "tractor_back": ("Тр.Об", "Тракторная оборотная"),
        "ambulatory_extract": ("ПЗ2",),
        "prof_ambulatory_extract": ("ПЗ2",),
        "prof_ambulatory": ("Амб !",),
        "070": ("CKK",),
        "072": ("CKK72",),
        "086": ("086",),
        "certificate_086": ("086",),
        "pool": ("Бас",),
        "sport": ("Спорт",),
        "gto": ("ГТО",),
        "gsu": ("ГС",),
        "gostaina": ("ГТ",),
        "gims": ("Суда",),
        "guard": ("ЧОД",),
        "chod": ("ЧОД",),
        "ekg": ("ЭЭГ",),
        "journal_344": ("Журн344",),
        "journal344": ("Журн344",),
        "spoiled": ("Испорч",),
    }
    target_sheet_names = sheets_by_variant.get(variant)
    if not target_sheet_names:
        raise ValueError(f"Неизвестный вариант печати: {print_variant}")

    if hasattr(target_book, "keep_only_sheets"):
        target_book.keep_only_sheets(target_sheet_names)
        return

    worksheets = list(getattr(target_book, "_Workbook__worksheets", []) or [])
    if not worksheets:
        return

    kept_sheet = next(
        (
            sheet
            for target_sheet_name in target_sheet_names
            for sheet in worksheets
            if getattr(sheet, "name", "") == target_sheet_name
        ),
        None,
    )
    if kept_sheet is None:
        raise ValueError(f"В шаблоне не найден лист для печати: {target_sheet_names[0]}")

    target_book._Workbook__worksheets = [kept_sheet]
    target_book._Workbook__worksheet_idx_from_name = {getattr(kept_sheet, "name", target_sheet_names[0]): 0}
    target_book._Workbook__active_sheet = 0


def _generate_xls(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    exams: list[DoctorExam],
) -> None:
    source_book = xlrd.open_workbook(file_contents=template_path.read_bytes(), formatting_info=True)
    if _find_prof_amb_sheet_index(source_book) is not None:
        _generate_prof_amb_xls(template_path, output_path, context, client, encounter, exams)
        return
    shutil.copy2(template_path, output_path)


def _new_xls_display_value(book, sheet, row_index: int, col_index: int) -> str:
    if row_index >= sheet.nrows or col_index >= sheet.ncols:
        return ""
    cell = sheet.cell(row_index, col_index)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        value = xldate.xldate_as_datetime(cell.value, book.datemode)
        return value.strftime("%d.%m.%Y")
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        number = float(cell.value)
        return str(int(number)) if number.is_integer() else str(number)
    return str(cell.value or "")


def _new_xls_workbook_stream(
    file_bytes: bytes,
) -> tuple[bytes, list[tuple[int, int, int]]]:
    compound_document = CompDoc(file_bytes)
    for stream_name in ("Workbook", "Book"):
        stream_bytes, stream_offset, stream_length = compound_document.locate_named_stream(stream_name)
        if stream_bytes is None or stream_length <= 0:
            continue
        directory_entry = compound_document._dir_search(stream_name.split("/"))
        if directory_entry is None or directory_entry.tot_size < compound_document.min_size_std_stream:
            raise ValueError(
                "Новый XLS-шаблон использует неподдерживаемый короткий Workbook stream"
            )
        logical_offset = 0
        sector_id = directory_entry.first_SID
        sectors: list[tuple[int, int, int]] = []
        while sector_id >= 0 and logical_offset < stream_length:
            sector_length = min(compound_document.sec_size, stream_length - logical_offset)
            physical_offset = (sector_id + 1) * compound_document.sec_size
            sectors.append(
                (
                    logical_offset,
                    logical_offset + sector_length,
                    physical_offset,
                )
            )
            logical_offset += sector_length
            sector_id = compound_document.SAT[sector_id]
        if logical_offset != stream_length:
            raise ValueError("Не удалось восстановить цепочку секторов Workbook stream")
        workbook_stream = bytes(stream_bytes[stream_offset : stream_offset + stream_length])
        return workbook_stream, sectors
    raise ValueError("В XLS-шаблоне не найден Workbook stream")


def _new_xls_sst_payload_spans(workbook_stream: bytes) -> list[tuple[int, int]]:
    spans: list[tuple[int, int]] = []
    offset = 0
    in_sst = False
    while offset + 4 <= len(workbook_stream):
        record_id, payload_length = struct.unpack_from("<HH", workbook_stream, offset)
        payload_start = offset + 4
        payload_end = payload_start + payload_length
        if payload_end > len(workbook_stream):
            break
        if record_id == 0x00FC:
            in_sst = True
            if payload_length >= 8:
                spans.append((payload_start + 8, payload_end))
        elif in_sst and record_id == 0x003C:
            spans.append((payload_start, payload_end))
        elif in_sst:
            break
        offset = payload_end
    if not spans:
        raise ValueError("В XLS-шаблоне не найдена таблица строк SST")
    return spans


def _new_xls_utf16_value(value: object) -> bytes:
    text = str(value or "").replace("\u200b", "").rstrip()
    while len(text.encode("utf-16le")) > PLACEHOLDER_LENGTH * 2:
        text = text[:-1]
    return text.encode("utf-16le")


def _write_new_xls_stream_bytes(
    file_bytes: bytearray,
    sectors: list[tuple[int, int, int]],
    logical_offset: int,
    value: bytes,
) -> None:
    remaining = value
    current_offset = logical_offset
    while remaining:
        sector = next(
            (
                item
                for item in sectors
                if item[0] <= current_offset < item[1]
            ),
            None,
        )
        if sector is None:
            raise ValueError("Запись вышла за пределы Workbook stream")
        logical_start, logical_end, physical_start = sector
        writable_length = min(len(remaining), logical_end - current_offset)
        target_start = physical_start + (current_offset - logical_start)
        file_bytes[target_start : target_start + writable_length] = remaining[:writable_length]
        remaining = remaining[writable_length:]
        current_offset += writable_length


def _new_xls_compact_sst_tail(
    encoded_values: list[bytes],
    *,
    first_capacity: int,
) -> tuple[bytes, int]:
    max_payload_length = 8224
    payloads: list[bytearray] = [bytearray()]
    capacities = [first_capacity]

    def start_payload(continuing_unicode_string: bool = False) -> None:
        payloads.append(bytearray(b"\x01" if continuing_unicode_string else b""))
        capacities.append(max_payload_length)

    for encoded_value in encoded_values:
        if capacities[-1] - len(payloads[-1]) < 3:
            start_payload()
        payloads[-1].extend(struct.pack("<HB", len(encoded_value) // 2, 0x01))
        character_offset = 0
        while character_offset < len(encoded_value):
            room = capacities[-1] - len(payloads[-1])
            writable = min(len(encoded_value) - character_offset, room)
            writable -= writable % 2
            if writable:
                payloads[-1].extend(
                    encoded_value[character_offset : character_offset + writable]
                )
                character_offset += writable
            if character_offset < len(encoded_value):
                start_payload(continuing_unicode_string=True)

    compacted = bytearray(payloads[0])
    for payload in payloads[1:]:
        compacted.extend(struct.pack("<HH", 0x003C, len(payload)))
        compacted.extend(payload)
    return bytes(compacted), len(payloads[0])


def _new_xls_patch_shifted_stream_offsets(
    workbook_stream: bytearray,
    *,
    effective_length: int,
    removed_start: int,
    removed_length: int,
) -> None:
    offset = 0
    while offset + 4 <= effective_length:
        record_id, payload_length = struct.unpack_from("<HH", workbook_stream, offset)
        payload_start = offset + 4
        payload_end = payload_start + payload_length
        if payload_end > effective_length:
            raise ValueError("Повреждена структура Workbook stream после уплотнения SST")
        if record_id == 0x0085 and payload_length >= 4:
            sheet_offset = struct.unpack_from("<I", workbook_stream, payload_start)[0]
            if sheet_offset >= removed_start:
                struct.pack_into(
                    "<I",
                    workbook_stream,
                    payload_start,
                    sheet_offset - removed_length,
                )
        elif record_id == 0x020B and payload_length >= 16:
            for pointer_offset in range(payload_start + 12, payload_end, 4):
                if pointer_offset + 4 > payload_end:
                    break
                dbcell_offset = struct.unpack_from("<I", workbook_stream, pointer_offset)[0]
                if dbcell_offset >= removed_start:
                    struct.pack_into(
                        "<I",
                        workbook_stream,
                        pointer_offset,
                        dbcell_offset - removed_length,
                    )
        elif record_id == 0x00FF:
            struct.pack_into("<H", workbook_stream, offset, 0x0000)
        offset = payload_end


def _patch_new_xls_placeholders(
    output_path: Path,
    spec: NewXlsTemplateSpec,
    values: dict[tuple[int, int], str],
) -> None:
    original_bytes = output_path.read_bytes()
    file_bytes = bytearray(original_bytes)
    workbook_stream, sectors = _new_xls_workbook_stream(original_bytes)
    spans = _new_xls_sst_payload_spans(workbook_stream)
    string_starts: list[int] = []
    for coordinate in spec.dynamic_cells:
        placeholder = new_xls_placeholder(spec, coordinate)
        marker = placeholder[:6].encode("utf-16le")
        marker_offset = workbook_stream.find(marker)
        if marker_offset < 0:
            raise ValueError(
                f"В {spec.file_name} не найден скрытый маркер ячейки "
                f"R{coordinate[0] + 1}C{coordinate[1] + 1}"
            )
        if workbook_stream.find(marker, marker_offset + 1) >= 0:
            raise ValueError(
                f"В {spec.file_name} скрытый маркер ячейки "
                f"R{coordinate[0] + 1}C{coordinate[1] + 1} не уникален"
            )

        span_index = next(
            (
                index
                for index, (span_start, span_end) in enumerate(spans)
                if span_start <= marker_offset < span_end
            ),
            None,
        )
        if span_index is None:
            raise ValueError("Скрытый маркер XLS находится вне SST payload")
        string_starts.append(marker_offset - 3)

    if string_starts != sorted(string_starts):
        raise ValueError("Скрытые поля XLS расположены не в порядке карты ячеек")
    first_string_start = string_starts[0]
    first_span_index = next(
        index
        for index, (span_start, span_end) in enumerate(spans)
        if span_start <= first_string_start < span_end
    )
    first_payload_start = (
        spans[first_span_index][0] - 8
        if first_span_index == 0
        else spans[first_span_index][0]
    )
    first_record_header = first_payload_start - 4
    first_payload_prefix_length = first_string_start - first_payload_start
    compacted_tail, first_tail_payload_length = _new_xls_compact_sst_tail(
        [
            _new_xls_utf16_value(values.get(coordinate, ""))
            for coordinate in spec.dynamic_cells
        ],
        first_capacity=8224 - first_payload_prefix_length,
    )
    old_tail_end = spans[-1][1]
    original_tail_length = old_tail_end - first_string_start
    removed_length = original_tail_length - len(compacted_tail)
    if removed_length < 0:
        raise ValueError("Данные не помещаются в область скрытых полей XLS")

    effective_length = len(workbook_stream) - removed_length
    compacted_stream = bytearray(
        workbook_stream[:first_string_start]
        + compacted_tail
        + workbook_stream[old_tail_end:]
        + b"\x00" * removed_length
    )
    struct.pack_into(
        "<H",
        compacted_stream,
        first_record_header + 2,
        first_payload_prefix_length + first_tail_payload_length,
    )
    _new_xls_patch_shifted_stream_offsets(
        compacted_stream,
        effective_length=effective_length,
        removed_start=old_tail_end,
        removed_length=removed_length,
    )
    _write_new_xls_stream_bytes(
        file_bytes,
        sectors,
        0,
        bytes(compacted_stream),
    )

    output_path.write_bytes(file_bytes)


def _generate_preserved_new_xls(
    template_path: Path,
    output_path: Path,
    source_book,
    spec: NewXlsTemplateSpec,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    runtime_values: dict[str, object],
    print_variant: str | None,
) -> None:
    target_book = copy_xls_workbook(source_book)
    exams_by_role = _exam_map(list(runtime_values.get("exams", [])))
    _fill_new_xls_sheets(
        source_book,
        target_book,
        context,
        client,
        encounter,
        exams_by_role,
    )
    _apply_print_variant_to_xls_workbook(target_book, print_variant or spec.print_variant)

    temporary_file = tempfile.NamedTemporaryFile(
        prefix=".new_xls_values_",
        suffix=".xls",
        dir=output_path.parent,
        delete=False,
    )
    temporary_path = Path(temporary_file.name)
    temporary_file.close()
    try:
        target_book.save(str(temporary_path))
        values_book = xlrd.open_workbook(
            file_contents=temporary_path.read_bytes(),
            formatting_info=True,
        )
        values_sheet = values_book.sheet_by_name(spec.sheet_name)
        values = {
            coordinate: _new_xls_display_value(
                values_book,
                values_sheet,
                coordinate[0],
                coordinate[1],
            )
            for coordinate in spec.dynamic_cells
        }
        shutil.copy2(template_path, output_path)
        _patch_new_xls_placeholders(output_path, spec, values)
    finally:
        temporary_path.unlink(missing_ok=True)


def _generate_runtime_xls(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    runtime_values: dict[str, object],
    print_variant: str | None = None,
) -> None:
    source_book = xlrd.open_workbook(file_contents=template_path.read_bytes(), formatting_info=True)
    new_template_spec = NEW_XLS_TEMPLATE_BY_FILE.get(template_path.name.casefold())
    if new_template_spec is not None:
        _generate_preserved_new_xls(
            template_path,
            output_path,
            source_book,
            new_template_spec,
            context,
            client,
            encounter,
            runtime_values,
            print_variant,
        )
        return
    exams = list(runtime_values.get("exams", []))
    if _find_prof_amb_sheet_index(source_book) is not None:
        _generate_prof_amb_xls(template_path, output_path, context, client, encounter, exams, print_variant=print_variant)
        return

    target_book = copy_xls_workbook(source_book)
    exams_by_role = _exam_map(exams)

    contract_source, contract_target, _ = _sheet_pair(source_book, target_book, "Договор !")
    if contract_source and contract_target:
        _fill_contract_xls_sheet(contract_source, contract_target, context, client, encounter, runtime_values)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "086")
    if source_sheet and target_sheet:
        _fill_086_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "ЭЭГ")
    if source_sheet and target_sheet:
        _fill_eeg_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "ЧОД")
    if source_sheet and target_sheet:
        _fill_chod_xls_sheet(source_sheet, target_sheet, context, encounter)

    _fill_driver_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)
    _fill_tractor_xls_sheets(source_book, target_book, exams_by_role)
    _fill_new_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "АмбОПО !")
    if source_sheet and target_sheet:
        _fill_amb_opo_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "Журн344")
    if source_sheet and target_sheet:
        _fill_journal_344_sheet(source_sheet, target_sheet, context, client, encounter)

    _apply_xls_auto_markers(source_book, target_book, context, client, encounter, exams_by_role)
    _apply_print_variant_to_xls_workbook(target_book, print_variant)
    target_book.save(str(output_path))


def _generate_runtime_xlsx(
    template_path: Path,
    output_path: Path,
    context: dict[str, str],
    client: Client,
    encounter: Encounter | None,
    runtime_values: dict[str, object],
    print_variant: str | None = None,
) -> None:
    source_book = _XlsxBookAdapter(load_workbook(template_path, data_only=False))
    exams = list(runtime_values.get("exams", []))
    if _find_prof_amb_sheet_index(source_book) is not None:
        _generate_prof_amb_xlsx(template_path, output_path, context, client, encounter, exams, print_variant=print_variant)
        return

    target_book = _XlsxBookAdapter(load_workbook(template_path, data_only=False))
    exams_by_role = _exam_map(exams)

    contract_source, contract_target, _ = _sheet_pair(source_book, target_book, "Договор !")
    if contract_source and contract_target:
        _fill_contract_xls_sheet(contract_source, contract_target, context, client, encounter, runtime_values)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "086")
    if source_sheet and target_sheet:
        _fill_086_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "ЭЭГ")
    if source_sheet and target_sheet:
        _fill_eeg_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "ЧОД")
    if source_sheet and target_sheet:
        _fill_chod_xls_sheet(source_sheet, target_sheet, context, encounter)

    _fill_driver_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)
    _fill_tractor_xls_sheets(source_book, target_book, exams_by_role)
    _fill_new_xls_sheets(source_book, target_book, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "АмбОПО !")
    if source_sheet and target_sheet:
        _fill_amb_opo_xls_sheet(source_sheet, target_sheet, context, client, encounter, exams_by_role)

    source_sheet, target_sheet, _ = _sheet_pair(source_book, target_book, "Журн344")
    if source_sheet and target_sheet:
        _fill_journal_344_sheet(source_sheet, target_sheet, context, client, encounter)

    _apply_xls_auto_markers(source_book, target_book, context, client, encounter, exams_by_role)
    _apply_print_variant_to_xls_workbook(target_book, print_variant)
    target_book.save(output_path)


def _first_field_value(fields: dict, *keys: str) -> str:
    lowered = {str(key).lower(): value for key, value in fields.items()}
    for key in keys:
        value = fields.get(key)
        if value in (None, ""):
            value = lowered.get(key.lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _parse_chairman_exam_date(value: object) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None

    ru_match = re.search(r"(?<!\d)(\d{1,2})[./](\d{1,2})[./](\d{2}|\d{4})(?!\d)", text)
    if ru_match:
        raw_date = ".".join(ru_match.groups())
        date_format = "%d.%m.%Y" if len(ru_match.group(3)) == 4 else "%d.%m.%y"
        try:
            return datetime.strptime(raw_date, date_format).date()
        except ValueError:
            return None

    iso_match = re.search(r"(?<!\d)(\d{4})-(\d{1,2})-(\d{1,2})(?!\d)", text)
    if iso_match:
        try:
            return date(*(int(part) for part in iso_match.groups()))
        except ValueError:
            return None

    return None


def _chairman_exam_date_context_overrides(exams: list[DoctorExam]) -> dict[str, str]:
    chairman = _exam_map(exams).get("chairman")
    if chairman is None:
        return {}

    exam_date = _parse_chairman_exam_date(
        _first_field_value(chairman.fields_json or {}, "examDate", "exam_date")
    )
    if exam_date is None:
        return {}

    formatted_date = exam_date.strftime("%d.%m.%y")
    day = exam_date.strftime("%d")
    month = exam_date.strftime("%m")
    month_name = MONTH_NAMES.get(exam_date.month, "")
    year = str(exam_date.year)
    valid_until = _add_calendar_months(exam_date, 6)
    return {
        "VisitDate": formatted_date,
        "VisitDate_DATEFULL": formatted_date,
        "DateCalc": formatted_date,
        "ServiceDateCalc": formatted_date,
        "ServiceDateCalc1": formatted_date,
        "VisitDate_DAY": day,
        "MonthCalc": month,
        "VisitDate_MONTH": month,
        "VisitDate_DATEMONTH": month_name,
        "VisitDate_YEAR": year,
        "VisitDate_DAY1": day,
        "MonthCalc1": month,
        "VisitDate_MONTH1": month,
        "VisitDate_DATEMONTH1": month_name,
        "VisitDate_YEAR1": year,
        "PoolValidUntil": valid_until.strftime("%d.%m.%Y"),
    }


def _chairman_certificate_date_context_overrides(
    template: DocumentTemplate,
    exams: list[DoctorExam],
) -> dict[str, str]:
    candidates = [getattr(template, "file_name", None), getattr(template, "file_path", None)]
    file_names = {
        Path(str(candidate)).name.casefold()
        for candidate in candidates
        if str(candidate or "").strip()
    }
    if not file_names.intersection(CHAIRMAN_EXAM_DATE_TEMPLATE_FILES):
        return {}
    return _chairman_exam_date_context_overrides(exams)


def _chairman_082_context_overrides(
    template: DocumentTemplate,
    exams: list[DoctorExam],
    blank_form: BlankForm | None,
) -> dict[str, str]:
    candidates = [getattr(template, "file_name", None), getattr(template, "file_path", None)]
    file_names = {
        Path(str(candidate)).name.casefold()
        for candidate in candidates
        if str(candidate or "").strip()
    }
    if not file_names.intersection(CHAIRMAN_CERTIFICATE_082_TEMPLATE_FILES):
        return {}

    chairman = _exam_map(exams).get("chairman")
    fields = (chairman.fields_json or {}) if chairman is not None else {}
    overrides = {
        "Country": _first_field_value(fields, "country", "destinationCountry", "destination_country"),
    }
    if blank_form is not None:
        full_number = str(blank_form.full_number or "").strip()
        series = str(blank_form.series or "").strip()
        if series and full_number.casefold().startswith(series.casefold()):
            sequence_number = full_number[len(series) :].strip()
        else:
            sequence_number = str(blank_form.number_value or "").strip()
        overrides["ReferenceNumber"] = sequence_number
        overrides["SeriesNumberCalc"] = sequence_number
    return overrides


def _sport_context_overrides(exams: list[DoctorExam]) -> dict[str, str]:
    empty_overrides = {
        "SportDiagnosis": "",
        "SportMedicalRequirements": "",
        "SportContraindications": "",
        "SportEkg": "",
        "SportEkgConclusion": "",
        "SportFluorography": "",
        "SportConclusionText": "",
        "SportConclusion": "",
        "ChairmanDoctor": "",
        "SportDoctor": "",
    }
    chairman = _exam_map(exams).get("chairman")
    if chairman is None:
        return empty_overrides

    fields = chairman.fields_json or {}
    diagnosis = _first_field_value(fields, "diagnosis", "diagnosisShort", "diagnosisText", "diagnoz")
    conclusion = _first_field_value(fields, "conclusion", "result")
    conclusion_text = _first_field_value(fields, "conclusionText", "sportConclusionText", "issuedConclusion")
    ekg = _first_field_value(fields, "ekg", "EKG", "ecg")
    ekg_conclusion = _first_field_value(fields, "ekgConclusion", "EKGConclusion", "ecgConclusion")
    medical_requirements = _first_field_value(fields, "medicalRequirements", "requirements")
    fluorography = _first_field_value(fields, "fluorography", "fluoro")
    chairman_doctor = str(chairman.doctor_name or "").strip()

    contraindications = ""
    if conclusion:
        contraindications = "выявлены" if conclusion.lower().startswith("не") else "не выявлены"

    return {
        **empty_overrides,
        "SportDiagnosis": diagnosis,
        "SportMedicalRequirements": medical_requirements,
        "SportContraindications": contraindications,
        "SportEkg": ekg,
        "SportEkgConclusion": ekg_conclusion,
        "SportFluorography": fluorography,
        "SportConclusionText": conclusion_text,
        "SportConclusion": _first_non_empty(conclusion_text, conclusion),
        "ChairmanDoctor": chairman_doctor,
        "SportDoctor": chairman_doctor,
    }


def _get_journal_info(template: DocumentTemplate) -> tuple[str, str] | None:
    name = f"{template.name} {template.file_name}".lower()
    if "вод" in name or "driver" in name or re.search(r"(?:^|\W)ву(?:$|\W)", name):
        return ("journal_344", "Журнал 344 водительских заключений")
    if "оруж" in name or "002" in name:
        return ("journal_441", "Журнал 441 оружейных заключений")
    if "лмк" in name:
        return ("lmk", "Журнал ЛМК")
    if "086" in name:
        return ("086", "Журнал справок 086у")
    return None


def _medical_record_context_overrides(medical_record: MedicalRecord | None) -> dict[str, str]:
    overrides: dict[str, str] = {
        "HealthGroup": "",
        "BloodGroup": "",
        "BloodType": "",
        "RhFactor": "",
        "RhesusFactor": "",
        "Allergies": "",
        "qdfMain.BloodGroup": "",
        "qdfMain.BloodType": "",
        "qdfMain.RhFactor": "",
        "qdfMain.Allergies": "",
        "gdfMain.BloodGroup": "",
        "gdfMain.BloodType": "",
        "gdfMain.RhFactor": "",
        "gdfMain.Allergies": "",
    }
    if medical_record is None:
        return overrides

    if medical_record.marital_status:
        overrides["MaritalStatus"] = medical_record.marital_status
    if medical_record.work_place:
        overrides["CompanyName"] = medical_record.work_place
        overrides["WorkPlace"] = medical_record.work_place
        overrides["qdfMain.WorkPlace"] = medical_record.work_place
    if medical_record.position:
        overrides["Post"] = medical_record.position
        overrides["PositionApplied"] = medical_record.position
        overrides["qdfMain.Post"] = medical_record.position
    blood_group = getattr(medical_record, "blood_group", None)
    rh_factor = getattr(medical_record, "rh_factor", None)
    allergies = getattr(medical_record, "allergies", None)
    if blood_group:
        overrides["BloodGroup"] = blood_group
        overrides["qdfMain.BloodGroup"] = blood_group
        overrides["gdfMain.BloodGroup"] = blood_group
    if rh_factor:
        overrides["RhFactor"] = rh_factor
        overrides["RhesusFactor"] = rh_factor
        overrides["qdfMain.RhFactor"] = rh_factor
        overrides["gdfMain.RhFactor"] = rh_factor
    if allergies:
        overrides["Allergies"] = allergies
        overrides["qdfMain.Allergies"] = allergies
        overrides["gdfMain.Allergies"] = allergies
    if medical_record.health_group:
        overrides["HealthGroup"] = medical_record.health_group
        overrides["BloodType"] = medical_record.health_group
        overrides["qdfMain.BloodType"] = medical_record.health_group
        overrides["gdfMain.BloodType"] = medical_record.health_group
    return overrides


def _apply_context_overrides(context: dict[str, str], overrides: dict[str, object] | None) -> None:
    context.update(
        {
            key: str(value).strip()
            for key, value in (overrides or {}).items()
            if value is not None
        }
    )


def _prof_29n_doctor_context_overrides(client: Client, exams: list[DoctorExam]) -> dict[str, str]:
    exams_by_role = _exam_map(exams)
    therapist = _exam_export_with_client_doctor(exams_by_role.get("therapist"), client, "therapist")
    chairman = _build_exam_export(exams_by_role.get("chairman"))
    psychiatrist = _exam_export_with_client_doctor(exams_by_role.get("psychiatrist"), client, "psychiatrist")
    psychiatrist_narcologist = _exam_export_with_client_doctor(
        exams_by_role.get("psychiatrist-narcologist"),
        client,
        "psychiatrist-narcologist",
    )
    commission_doctor = _first_non_empty(
        therapist.get("doctor"),
        psychiatrist.get("doctor"),
        psychiatrist_narcologist.get("doctor"),
        chairman.get("doctor"),
    )
    return {
        "Prof29ChairmanDoctor": commission_doctor,
        "Prof29PathologistDoctor": commission_doctor,
        "Prof29PsychiatristDoctor": str(psychiatrist.get("doctor") or "").strip(),
        "Prof29PsychiatristNarcologistDoctor": str(psychiatrist_narcologist.get("doctor") or "").strip(),
    }


def _document_doctor_name_for_context(exams: list[DoctorExam]) -> str:
    doctor_names: list[str] = []
    chairman_doctor = ""
    for exam in exams:
        doctor_name = str(exam.doctor_name or "").strip()
        if not doctor_name:
            continue
        if doctor_name not in doctor_names:
            doctor_names.append(doctor_name)
        if str(exam.doctor_role_id or "") == "chairman" and not chairman_doctor:
            chairman_doctor = doctor_name
    return chairman_doctor or ", ".join(doctor_names)


def _load_encounter_document_values(db: Session, client: Client, encounter: Encounter | None) -> dict[str, object]:
    if encounter is None:
        fallback_services = client.legacy_payload_json.get("services", []) if isinstance(client.legacy_payload_json, dict) else []
        service_names = [str(item).strip() for item in fallback_services if str(item).strip()]
        service_rows = [
            {
                "ordinal": str(index),
                "service": name,
                "quantity": "1",
                "date": date.today().strftime("%d.%m.%y"),
                "unit_price": "",
                "line_total": "",
            }
            for index, name in enumerate(service_names, start=1)
        ]
        medical_record = db.execute(
            select(MedicalRecord)
            .where(MedicalRecord.client_id == client.id, MedicalRecord.deleted_at.is_(None))
            .order_by(MedicalRecord.updated_at.desc(), MedicalRecord.id.desc())
        ).scalars().first()
        context_overrides = _medical_record_context_overrides(medical_record)
        context_overrides.update(_prof_29n_doctor_context_overrides(client, []))
        return {
            "service_names": service_names,
            "service_rows": service_rows,
            "doctor_name": "",
            "diagnosis": "",
            "mkb10": "",
            "exams": [],
            "context_overrides": context_overrides,
        }

    service_items = (
        db.execute(
            select(EncounterService, Service.name)
            .join(Service, EncounterService.service_id == Service.id)
            .where(EncounterService.encounter_id == encounter.id)
            .order_by(EncounterService.id.asc())
        )
        .all()
    )
    service_names = [name for _, name in service_items]
    service_rows = [
        {
            "ordinal": str(index),
            "service": name,
            "quantity": str(item.quantity or 1),
            "date": encounter.encounter_date.strftime("%d.%m.%y"),
            "unit_price": str(item.unit_price or ""),
            "line_total": str(item.line_total or ""),
        }
        for index, (item, name) in enumerate(service_items, start=1)
    ]
    if not service_rows:
        fallback_services = client.legacy_payload_json.get("services", []) if isinstance(client.legacy_payload_json, dict) else []
        service_names = [str(item).strip() for item in fallback_services if str(item).strip()]
        service_rows = [
            {
                "ordinal": str(index),
                "service": name,
                "quantity": "1",
                "date": encounter.encounter_date.strftime("%d.%m.%y"),
                "unit_price": "",
                "line_total": "",
            }
            for index, name in enumerate(service_names, start=1)
        ]

    exams = (
        db.execute(
            select(DoctorExam)
            .where(DoctorExam.encounter_id == encounter.id, DoctorExam.deleted_at.is_(None))
            .order_by(DoctorExam.updated_at.desc(), DoctorExam.id.desc())
        )
        .scalars()
        .all()
    )
    diagnosis = ""
    mkb10 = ""
    medical_record = db.execute(
        select(MedicalRecord)
        .where(MedicalRecord.client_id == client.id, MedicalRecord.deleted_at.is_(None))
        .order_by(MedicalRecord.updated_at.desc(), MedicalRecord.id.desc())
    ).scalars().first()
    context_overrides = _medical_record_context_overrides(medical_record)
    context_overrides.update(
        {
            "Weight": "",
            "Height": "",
            "HairColor": "",
            "EyeColor": "",
            "DistinguishingMark": "",
        }
    )
    for exam in exams:
        fields = exam.fields_json or {}
        diagnosis = diagnosis or _first_field_value(
            fields,
            "diagnosis",
            "diagnosisShort",
            "diagnosisText",
            "diagnose",
            "diagnoz",
            "conclusion",
        )
        mkb10 = mkb10 or _first_field_value(fields, "mkb10", "mkb", "icd10")
        context_overrides["Weight"] = context_overrides["Weight"] or _first_field_value(fields, "weight", "Weight")
        context_overrides["Height"] = context_overrides["Height"] or _first_field_value(fields, "height", "Height")
        context_overrides["HairColor"] = context_overrides["HairColor"] or _first_field_value(
            fields, "hairColor", "hair", "hair_color"
        )
        context_overrides["EyeColor"] = context_overrides["EyeColor"] or _first_field_value(
            fields, "eyeColor", "eyesColor", "eye_color"
        )
        context_overrides["DistinguishingMark"] = context_overrides["DistinguishingMark"] or _first_field_value(
            fields, "distinguishingMark", "distinguishingMarks", "specialMarks", "special_mark"
        )
    context_overrides.update(
        {
            key: value
            for key, value in _driver_document_context_overrides(client, exams).items()
            if value not in (None, "")
        }
    )
    context_overrides.update(
        {
            key: value
            for key, value in _sport_context_overrides(exams).items()
            if value not in (None, "")
        }
    )
    context_overrides.update(_prof_29n_doctor_context_overrides(client, exams))

    return {
        "service_names": service_names,
        "service_rows": service_rows,
        "doctor_name": _document_doctor_name_for_context(exams),
        "diagnosis": diagnosis,
        "mkb10": mkb10,
        "exams": exams,
        "context_overrides": context_overrides,
    }


def _append_blank_entry_to_medical_record_legacy(
    db: Session,
    *,
    client: Client,
    encounter: Encounter,
    blank_number: str,
) -> None:
    medical_record = db.execute(
        select(MedicalRecord).where(MedicalRecord.client_id == client.id, MedicalRecord.deleted_at.is_(None))
    ).scalar_one_or_none()
    if medical_record is None:
        medical_record = MedicalRecord(
            client_id=client.id,
            center_id=encounter.center_id,
            card_number=client.card_number,
            opened_at=encounter.encounter_date,
            oms_policy=client.oms_policy,
            work_place=client.work_place,
            position=client.profession,
            mkb10=client.mkb10,
            notes=client.notes,
        )
        db.add(medical_record)
        db.flush()

    db.add(
        MedicalRecordEntry(
            medical_record_id=medical_record.id,
            encounter_id=encounter.id,
            entry_date=encounter.encounter_date,
            doctor_role_id="document",
            doctor_name="document",
            conclusion=f"Выдан номерной бланк медицинского заключения №{blank_number}",
        )
    )


def _append_blank_entry_to_medical_record(
    db: Session,
    *,
    client: Client,
    encounter: Encounter,
    blank_number: str,
) -> None:
    conclusion = f"Выдан номерной бланк медицинского заключения №{blank_number}"
    medical_record = db.execute(
        select(MedicalRecord).where(MedicalRecord.client_id == client.id, MedicalRecord.deleted_at.is_(None))
    ).scalar_one_or_none()
    if medical_record is None:
        medical_record = MedicalRecord(
            client_id=client.id,
            center_id=encounter.center_id,
            card_number=client.card_number,
            opened_at=encounter.encounter_date,
            oms_policy=client.oms_policy,
            work_place=client.work_place,
            position=client.profession,
            mkb10=client.mkb10,
            notes=client.notes,
        )
        db.add(medical_record)
        db.flush()

    existing_entry = db.execute(
        select(MedicalRecordEntry).where(
            MedicalRecordEntry.medical_record_id == medical_record.id,
            MedicalRecordEntry.encounter_id == encounter.id,
            MedicalRecordEntry.doctor_role_id == "document",
            MedicalRecordEntry.conclusion == conclusion,
        )
    ).scalar_one_or_none()
    if existing_entry is not None:
        return

    db.add(
        MedicalRecordEntry(
            medical_record_id=medical_record.id,
            encounter_id=encounter.id,
            entry_date=encounter.encounter_date,
            doctor_role_id="document",
            doctor_name="document",
            conclusion=conclusion,
        )
    )


def generate_document(
    db: Session,
    *,
    template_id: int | None,
    template_code: str | None,
    client_id: int,
    encounter_id: int | None,
    blank_form_id: int | None = None,
    print_variant: str | None = None,
) -> DocumentGenerateResponse:
    print_variant_value = str(print_variant or "").strip().lower()
    side_print_variants = {"driver_front", "driver_back", "tractor_front", "tractor_back"}
    is_side_print = print_variant_value in side_print_variants
    template = None
    if template_id is not None:
        template = db.get(DocumentTemplate, template_id)
    elif template_code is not None:
        template = db.execute(select(DocumentTemplate).where(DocumentTemplate.code == template_code)).scalar_one_or_none()

    if template is None or not template.file_path:
        raise ValueError("Шаблон не найден")

    miac_xml_kind = _miac_xml_kind(template)

    client = db.get(Client, client_id)
    if client is None or client.deleted_at is not None:
        raise ValueError("Клиент не найден")

    encounter = None
    if encounter_id is not None:
        encounter = db.get(Encounter, encounter_id)
        if encounter is None or encounter.deleted_at is not None:
            raise ValueError("Обращение не найдено")

    template_path = Path(template.file_path)
    output_dir = Path(settings.generated_documents_dir)
    if template.template_type == "xml":
        output_dir = output_dir / "xml" / _xml_export_date_folder()
    elif _is_contract_template(template):
        output_dir = output_dir / "contracts"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file_name = f"{template_path.stem}_{client.id}_{timestamp}{template_path.suffix}"
    output_path = output_dir / output_file_name

    runtime_values = _load_encounter_document_values(db, client, encounter)
    required_blank_type = (
        None
        if is_side_print
        else resolve_required_blank_type(template, print_variant=print_variant_value)
    )
    blank_form = None

    try:
        if required_blank_type:
            if encounter is None or encounter.center_id is None:
                raise ValueError(
                    "Для документа с номерным бланком требуется encounter_id и center_id. "
                    "Сначала оформите обращение в нужном медцентре."
                )

            if miac_xml_kind is not None:
                candidate_form = db.get(BlankForm, blank_form_id) if blank_form_id is not None else None
                if miac_xml_kind == "gims" and candidate_form is not None and candidate_form.status == BLANK_STATUS_FREE:
                    blank_form = issue_specific_blank(
                        db,
                        form_id=candidate_form.id,
                        blank_type=required_blank_type,
                        client_id=client.id,
                        center_id=encounter.center_id,
                        encounter_id=encounter.id,
                        user_id=1,
                    )
                else:
                    blank_form = _resolve_miac_issued_blank(
                        db,
                        blank_type=required_blank_type,
                        client_id=client.id,
                        encounter_id=encounter.id,
                        center_id=encounter.center_id,
                        blank_form_id=blank_form_id,
                    )
            else:
                blank_form = reuse_blank_for_existing_document(
                    db,
                    blank_type=required_blank_type,
                    client_id=client.id,
                    encounter_id=encounter.id,
                    template_id=template.id,
                )
                if blank_form is None:
                    if blank_form_id is not None:
                        blank_form = issue_specific_blank(
                            db,
                            form_id=blank_form_id,
                            blank_type=required_blank_type,
                            client_id=client.id,
                            center_id=encounter.center_id,
                            encounter_id=encounter.id,
                            user_id=1,
                        )
                    else:
                        blank_form = issue_next_blank(
                            db,
                            blank_type=required_blank_type,
                            client_id=client.id,
                            center_id=encounter.center_id,
                            encounter_id=encounter.id,
                            user_id=1,
                        )

        elif blank_form_id is not None and not is_side_print:
            if encounter is None or encounter.center_id is None:
                raise ValueError(
                    "Р”Р»СЏ РїРµС‡Р°С‚Рё РЅР° РЅРѕРјРµСЂРЅРѕРј Р±Р»Р°РЅРєРµ С‚СЂРµР±СѓРµС‚СЃСЏ encounter_id Рё center_id. "
                    "РЎРЅР°С‡Р°Р»Р° РѕС„РѕСЂРјРёС‚Рµ РѕР±СЂР°С‰РµРЅРёРµ РІ РЅСѓР¶РЅРѕРј РјРµРґС†РµРЅС‚СЂРµ."
                )
            candidate_form = db.get(BlankForm, blank_form_id)
            if candidate_form is None:
                raise ValueError("Р‘Р»Р°РЅРє РЅРµ РЅР°Р№РґРµРЅ")
            blank_type = template.blank_type or candidate_form.blank_type or BLANK_TYPE_DRIVER_MEDICAL_CERTIFICATE
            if candidate_form.status == BLANK_STATUS_FREE:
                blank_form = issue_specific_blank(
                    db,
                    form_id=blank_form_id,
                    blank_type=blank_type,
                    client_id=client.id,
                    center_id=encounter.center_id,
                    encounter_id=encounter.id,
                    user_id=1,
                )
            elif (
                candidate_form.status == BLANK_STATUS_ISSUED
                and candidate_form.client_id == client.id
                and candidate_form.encounter_id == encounter.id
            ):
                blank_form = candidate_form
            else:
                raise ValueError("Р’С‹Р±СЂР°РЅРЅС‹Р№ Р±Р»Р°РЅРє СѓР¶Рµ РЅРµРґРѕСЃС‚СѓРїРµРЅ РґР»СЏ СЌС‚РѕР№ РїРµС‡Р°С‚Рё")

        if is_side_print and blank_form_id is not None:
            if encounter is None or encounter.center_id is None:
                raise ValueError(
                    "Для печати на номерном бланке требуется encounter_id и center_id. "
                    "Сначала оформите обращение в нужном медцентре."
                )
            candidate_form = db.get(BlankForm, blank_form_id)
            if candidate_form is None:
                raise ValueError("Бланк не найден")
            if candidate_form.status == BLANK_STATUS_FREE:
                blank_form = issue_specific_blank(
                    db,
                    form_id=blank_form_id,
                    blank_type=template.blank_type or BLANK_TYPE_DRIVER_MEDICAL_CERTIFICATE,
                    client_id=client.id,
                    center_id=encounter.center_id,
                    encounter_id=encounter.id,
                    user_id=1,
                )
            elif (
                candidate_form.status == BLANK_STATUS_ISSUED
                and candidate_form.client_id == client.id
                and candidate_form.encounter_id == encounter.id
            ):
                blank_form = candidate_form
            else:
                raise ValueError("Выбранный бланк уже недоступен для этой печати")

        context = build_document_context(
            client,
            encounter,
            service_names=runtime_values["service_names"],
            doctor_name=runtime_values["doctor_name"],
            diagnosis=runtime_values["diagnosis"],
            mkb10=runtime_values["mkb10"],
        )
        _apply_context_overrides(context, runtime_values.get("context_overrides"))
        document_exams = list(runtime_values.get("exams", []))
        context.update(_driver_document_context_overrides(client, document_exams))
        context.update(
            {
                key: str(value or "").strip()
                for key, value in _sport_context_overrides(document_exams).items()
            }
        )
        context.update(_chairman_certificate_date_context_overrides(template, document_exams))
        if blank_form is not None:
            context["BlankNumber"] = blank_form.full_number
            context["BlankSeries"] = blank_form.series or ""
            context["BlankFullNumber"] = blank_form.full_number
            context["DocumentNumber"] = blank_form.full_number
        context.update(_chairman_082_context_overrides(template, document_exams, blank_form))
        if is_side_print:
            context["ReferenceNumber"] = ""
            context["SeriesNumberCalc"] = ""
            context["BlankNumber"] = ""
            context["BlankSeries"] = ""
            context["BlankFullNumber"] = ""

        if template.template_type == "docx":
            _generate_docx(
                template_path,
                output_path,
                context,
                runtime_values["service_rows"],
                cleanup_xml=_is_contract_template(template),
            )
        elif template.template_type == "xml" and miac_xml_kind is not None:
            _generate_miac_xml(
                output_path,
                kind=miac_xml_kind,
                client=client,
                blank_form=blank_form,
                exams=document_exams,
            )
        elif template.template_type == "xml":
            xml_context = context.copy()
            xml_context.update(_driver_xml_context_overrides(xml_context, client, document_exams))
            _generate_xml(template_path, output_path, xml_context)
        elif template.template_type == "xls":
            _generate_runtime_xls(
                template_path,
                output_path,
                context,
                client,
                encounter,
                runtime_values,
                print_variant=print_variant,
            )
        elif template.template_type == "xlsx":
            _generate_runtime_xlsx(
                template_path,
                output_path,
                context,
                client,
                encounter,
                runtime_values,
                print_variant=print_variant,
            )
        else:
            shutil.copy2(template_path, output_path)

        document_number = blank_form.full_number if blank_form is not None else client.reference_number
        document_series = blank_form.series if blank_form is not None else client.document_series

        generated_document = GeneratedDocument(
            encounter_id=encounter.id if encounter else None,
            client_id=client.id,
            template_id=template.id,
            document_number=document_number,
            series=document_series,
            file_name=output_file_name,
            file_path=str(output_path.resolve()),
            generated_by_user_id=1,
            blank_form_id=blank_form.id if blank_form is not None else None,
            blank_number_snapshot=blank_form.full_number if blank_form is not None else None,
        )
        db.add(generated_document)
        db.flush()

        if blank_form is not None and blank_form.generated_document_id is None:
            blank_form.generated_document_id = generated_document.id
            db.flush()

        journal_info = _get_journal_info(template)
        if journal_info is not None:
            journal_code, journal_name = journal_info
            db.add(
                DocumentJournalEntry(
                    journal_code=journal_code,
                    journal_name=journal_name,
                    generated_document_id=generated_document.id,
                    client_id=client.id,
                    encounter_id=encounter.id if encounter else None,
                    issued_at=encounter.encounter_date if encounter else None,
                    series=generated_document.series,
                    number=generated_document.document_number,
                    result_text=context.get("Diagnosis") or context.get("Conclusion") or "",
                    created_by_user_id=1,
                )
            )

        if blank_form is not None and encounter is not None:
            _append_blank_entry_to_medical_record(
                db,
                client=client,
                encounter=encounter,
                blank_number=blank_form.full_number,
            )

        write_audit_log(
            db,
            entity_type="document_template",
            entity_id=template.id,
            action="generate",
            user_id=1,
            center_id=encounter.center_id if encounter else None,
            payload_json={
                "client_id": client.id,
                "encounter_id": encounter.id if encounter else None,
                "blank_form_id": blank_form.id if blank_form is not None else None,
                "blank_number": blank_form.full_number if blank_form is not None else None,
            },
        )

        return DocumentGenerateResponse(
            template_name=template.name,
            template_type=template.template_type,
            output_file_name=output_file_name,
            output_file_path=str(output_path.resolve()),
            generated_document_id=generated_document.id,
            blank_form_id=blank_form.id if blank_form is not None else None,
            blank_number=blank_form.full_number if blank_form is not None else None,
            generated_fields=context,
        )
    except Exception:
        if output_path.exists():
            output_path.unlink(missing_ok=True)
        raise
