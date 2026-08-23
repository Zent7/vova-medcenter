"""Генерация Excel-шаблона для загрузки клиентов.

Шаблон намеренно не содержит справочников из базы (услуг, цен): такой файл
живёт у заказчика неделями, и любой снимок прайса в нём протухает.
Услугу и обращение медцентр назначает уже внутри системы.

Запуск:
    python backend/scripts/generate_client_import_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SHEET_NAME = "Клиенты"
HELP_SHEET_NAME = "Инструкция"
DATA_ROWS = 300

HEADER_FILL = "FF0F766E"
REQUIRED_FILL = "FFFFF7CC"
PLAIN_FILL = "FFFFFFFF"
TEXT_COLOR = "FF1F2937"
BORDER_COLOR = "FFD1D5DB"
FONT_NAME = "Calibri"

# (заголовок, ширина, обязательная, числовой формат)
COLUMNS = [
    ("№ пациента", 14, False, "0"),
    ("Фамилия", 18, True, "@"),
    ("Имя", 18, True, "@"),
    ("Отчество", 20, False, "@"),
    ("Дата рождения", 17, True, "dd.mm.yyyy"),
    ("Пол", 11, False, "@"),
    ("Телефон", 19, False, "@"),
    ("СНИЛС", 20, False, "@"),
    ("Регистрация", 34, False, "@"),
    ("Организация", 28, False, "@"),
    ("Примечание", 34, False, "@"),
]

HELP_STEPS = [
    ("1", "Заполняйте только первый лист «Клиенты». Одна строка — один человек."),
    ("2", "Жёлтые колонки обязательны: фамилия, имя, дата рождения."),
    ("3", "Дату рождения пишите как 15.04.1987."),
    ("4", "Пол выбирайте из выпадающего списка: муж или жен."),
    ("5", "Колонку «№ пациента» оставьте пустой — номер присвоит система."),
    ("6", "Телефон и СНИЛС можно писать в любом виде, они сохранятся как текст."),
    ("7", "Готовый файл загрузите на странице «Загрузка клиентов»: сначала предпросмотр, потом загрузка в базу."),
]

THIN_BORDER = Border(*(Side(style="thin", color=BORDER_COLOR),) * 4)


def build_clients_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 34

    for index, (header, width, required, number_format) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = width

        cell = worksheet.cell(row=1, column=index, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

        fill = PatternFill("solid", fgColor=REQUIRED_FILL if required else PLAIN_FILL)
        for row_number in range(2, DATA_ROWS + 1):
            data_cell = worksheet.cell(row=row_number, column=index)
            data_cell.font = Font(name=FONT_NAME, size=10, color=TEXT_COLOR)
            data_cell.fill = fill
            data_cell.border = THIN_BORDER
            data_cell.number_format = number_format

    sex_column = get_column_letter(next(i for i, item in enumerate(COLUMNS, start=1) if item[0] == "Пол"))
    sex_validation = DataValidation(type="list", formula1='"муж,жен"', allow_blank=True, showErrorMessage=True)
    worksheet.add_data_validation(sex_validation)
    sex_validation.add(f"{sex_column}2:{sex_column}{DATA_ROWS}")

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{DATA_ROWS}"
    worksheet.freeze_panes = "A2"


def build_help_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 96
    worksheet.row_dimensions[1].height = 34

    title = worksheet.cell(row=1, column=1, value="Как заполнить файл")
    title.font = Font(name=FONT_NAME, bold=True, size=16, color="FFFFFFFF")
    title.fill = PatternFill("solid", fgColor=HEADER_FILL)
    title.alignment = Alignment(vertical="center")
    for column in range(2, 3):
        worksheet.cell(row=1, column=column).fill = PatternFill("solid", fgColor=HEADER_FILL)

    for column, header in ((1, "Шаг"), (2, "Что сделать")):
        cell = worksheet.cell(row=3, column=column, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF1D4ED8")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for offset, (step, text) in enumerate(HELP_STEPS):
        row_number = 4 + offset
        step_cell = worksheet.cell(row=row_number, column=1, value=step)
        step_cell.font = Font(name=FONT_NAME, bold=True, size=11, color="FF1E3A8A")
        step_cell.fill = PatternFill("solid", fgColor="FFDBEAFE")
        step_cell.alignment = Alignment(horizontal="center", vertical="center")

        text_cell = worksheet.cell(row=row_number, column=2, value=text)
        text_cell.font = Font(name=FONT_NAME, size=11, color=TEXT_COLOR)
        text_cell.alignment = Alignment(vertical="center")

    note_row = 4 + len(HELP_STEPS) + 1
    note = worksheet.cell(
        row=note_row,
        column=1,
        value="Услугу и дату приёма назначает медцентр после загрузки — в файле их указывать не нужно.",
    )
    note.font = Font(name=FONT_NAME, italic=True, size=11, color=TEXT_COLOR)


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    outputs = [
        root / "frontend" / "public" / "demo" / "client-import-template.xlsx",
        root / "frontend" / "dist" / "demo" / "client-import-template.xlsx",
    ]

    workbook = Workbook()
    build_clients_sheet(workbook.active)
    workbook.active.title = SHEET_NAME
    build_help_sheet(workbook.create_sheet(HELP_SHEET_NAME))

    for output in outputs:
        if not output.parent.exists():
            continue
        workbook.save(output)
        print(output)


if __name__ == "__main__":
    main()
