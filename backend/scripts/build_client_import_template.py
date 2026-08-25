"""Собирает Excel-шаблон, который завод заполняет для загрузки клиентов.

Файл кладётся в frontend/public/demo/client-import-template.xlsx — его
скачивают кнопкой «Скачать шаблон Excel» в разделе «Загрузка клиентов».
Заголовки должны совпадать с CLIENT_IMPORT_HEADERS в
backend/app/api/v1/routes/imports.py, иначе колонка приедет пустой.

Запуск: python backend/scripts/build_client_import_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_ROWS = 300
HEADER_FILL = PatternFill("solid", fgColor="FF0C7D7B")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
REQUIRED_FILL = PatternFill("solid", fgColor="FFFFF2CC")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True)

# (заголовок, ширина, обязательная ли колонка)
COLUMNS = [
    ("Тип документа", 19, False),
    ("Фамилия", 18, True),
    ("Имя", 18, True),
    ("Отчество", 20, False),
    ("Дата рождения (формат 31.12.2026)", 34, True),
    ("Пол", 13, False),
    ("Адрес Регистрация-ОБЛАСТЬ", 32, False),
    ("Адрес Регистрация-ГОРОД", 29, False),
    ("Адрес Регистрация-УЛИЦА", 27, False),
    ("Адрес Регистрация-НОМЕР ДОМА", 32, False),
    ("корпус, литер, строение", 28, False),
    ("квартира", 14, False),
    ("Название Организация", 26, False),
    ("Должность", 22, False),
    ("Вредные произв. Факторы", 26, False),
    ("СНИЛС", 20, False),
    ("ФЛГ от", 14, False),
    ("Примечание", 30, False),
]

SERVICE_VALUES = ["ЛМК", "Проф", "Водительская", "Тракторная", "ГИМС", "Продление ЛМК"]

INSTRUCTIONS = [
    "Заполняйте только первый лист «Clients». Одна строка — один человек.",
    "Жёлтые колонки обязательны: фамилия, имя, дата рождения.",
    "Дату рождения пишите как 15.04.1987.",
    "«Тип документа» и «Пол» выбирайте из выпадающего списка в ячейке.",
    "Адрес регистрации разнесите по колонкам: область, город, улица, дом, корпус, квартира.",
    "«ФЛГ от» — дата последней флюорографии, тоже в виде 15.04.1987.",
    "Пустые колонки можно не заполнять: медцентр дозаполнит их сам.",
    "Готовый файл отправьте в медцентр — там его загрузят в программу.",
]


def build_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients"

    for index, (header, width, required) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[letter].width = width
        for row in range(2, DATA_ROWS + 2):
            data_cell = sheet.cell(row=row, column=index)
            # Текстовый формат: дата и СНИЛС должны сохраниться так, как их
            # набрали, а не превратиться в число вида 31882.
            data_cell.number_format = "@"
            if required:
                data_cell.fill = REQUIRED_FILL

    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    service_validation = DataValidation(
        type="list", formula1='"' + ",".join(SERVICE_VALUES) + '"', allow_blank=True
    )
    sheet.add_data_validation(service_validation)
    service_validation.add(f"A2:A{DATA_ROWS + 1}")

    sex_validation = DataValidation(type="list", formula1='"муж,жен"', allow_blank=True)
    sheet.add_data_validation(sex_validation)
    sex_validation.add(f"F2:F{DATA_ROWS + 1}")

    guide = workbook.create_sheet("Инструкция")
    guide.column_dimensions["A"].width = 8
    guide.column_dimensions["B"].width = 96
    guide.sheet_view.showGridLines = False
    guide["A1"] = "Как заполнять шаблон"
    guide["A1"].font = TITLE_FONT
    for offset, text in enumerate(INSTRUCTIONS):
        guide.cell(row=3 + offset, column=1, value=f"{offset + 1}.")
        guide.cell(row=3 + offset, column=2, value=text)

    list_row = 4 + len(INSTRUCTIONS)
    guide.cell(row=list_row, column=1, value="Тип документа").font = SUBTITLE_FONT
    guide.cell(row=list_row, column=2, value=", ".join(SERVICE_VALUES))

    return workbook


def main() -> None:
    target = Path(__file__).resolve().parents[2] / "frontend" / "public" / "demo" / "client-import-template.xlsx"
    build_workbook().save(target)
    print(f"Шаблон сохранён: {target}")


if __name__ == "__main__":
    main()
